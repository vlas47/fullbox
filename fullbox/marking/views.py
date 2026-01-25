import json

from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Count
from django.http import HttpResponseBadRequest, HttpResponseForbidden, JsonResponse
from django.views.decorators.http import require_GET, require_POST
from openpyxl import load_workbook

from audit.models import OrderAuditEntry
from employees.access import get_request_role
from sku.models import SKU, SKUBarcode
from .models import MarkingCode
from .utils import extract_processing_items

ALLOWED_PROCESSING_ROLES = {"storekeeper", "processing_head", "head_manager", "director", "admin"}


def _parse_json_body(request):
    if not request.body:
        return {}
    try:
        return json.loads(request.body.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        return None


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text


def _get_processing_order(order_id: str):
    if not order_id:
        return None, None, None
    latest = (
        OrderAuditEntry.objects.filter(order_id=order_id, order_type="processing")
        .select_related("agency")
        .order_by("-created_at")
        .first()
    )
    if not latest:
        return None, None, None
    return latest, latest.payload or {}, latest.agency


def _resolve_sku(agency, sku_code: str):
    if not sku_code:
        return None
    qs = SKU.objects.filter(sku_code=sku_code)
    if agency:
        sku = qs.filter(agency=agency).first()
        if sku:
            return sku
        return qs.filter(agency__isnull=True).first()
    return qs.first()


def _require_processing_role(request):
    role = get_request_role(request)
    if role not in ALLOWED_PROCESSING_ROLES:
        return False, HttpResponseForbidden("Доступ запрещен")
    return True, None


@login_required
@require_GET
def processing_marking_summary(request, order_id: str):
    ok, response = _require_processing_role(request)
    if not ok:
        return response
    latest, payload, _agency = _get_processing_order(order_id)
    if not latest:
        return HttpResponseBadRequest("Заявка не найдена")
    rows = (
        MarkingCode.objects.filter(order_type="processing", order_id=order_id)
        .values("sku_code", "size")
        .annotate(count=Count("id"))
    )
    items = [
        {"sku_code": row["sku_code"], "size": row["size"] or "", "count": row["count"]}
        for row in rows
    ]
    total_count = sum(row["count"] for row in rows)
    return JsonResponse({"ok": True, "items": items, "total_count": total_count})


@login_required
@require_POST
def processing_marking_scan(request, order_id: str):
    ok, response = _require_processing_role(request)
    if not ok:
        return response
    latest, payload, agency = _get_processing_order(order_id)
    if not latest:
        return HttpResponseBadRequest("Заявка не найдена")
    data = _parse_json_body(request)
    if data is None:
        return HttpResponseBadRequest("Некорректный JSON")
    code = (data.get("code") or "").strip()
    sku_code = (data.get("sku_code") or "").strip()
    size = (data.get("size") or "").strip()
    barcode = (data.get("barcode") or "").strip()
    if not code:
        return JsonResponse({"ok": False, "error": "Код ЧЗ не указан"}, status=400)
    if not sku_code:
        return JsonResponse({"ok": False, "error": "Артикул не указан"}, status=400)
    items = extract_processing_items(payload)
    allowed_pairs = {(item["sku_code"], item["size"]) for item in items}
    if not size:
        matched = [pair for pair in allowed_pairs if pair[0] == sku_code]
        if len(matched) == 1:
            size = matched[0][1]
    if (sku_code, size) not in allowed_pairs:
        if size and (sku_code, "") in allowed_pairs:
            size = ""
        else:
            return JsonResponse(
                {"ok": False, "error": "Позиция не найдена в заявке."},
                status=400,
            )
    if MarkingCode.objects.filter(code=code).exists():
        return JsonResponse({"ok": False, "error": "Код уже учтен."}, status=409)
    sku = _resolve_sku(agency, sku_code)
    try:
        MarkingCode.objects.create(
            order_type="processing",
            order_id=order_id,
            agency=agency,
            sku=sku,
            sku_code=sku_code,
            size=size,
            barcode=barcode,
            code=code,
            source="scan",
            created_by=request.user if request.user.is_authenticated else None,
        )
    except IntegrityError:
        return JsonResponse({"ok": False, "error": "Код уже учтен."}, status=409)
    count = MarkingCode.objects.filter(
        order_type="processing",
        order_id=order_id,
        sku_code=sku_code,
        size=size,
    ).count()
    total_count = MarkingCode.objects.filter(
        order_type="processing",
        order_id=order_id,
    ).count()
    return JsonResponse(
        {
            "ok": True,
            "sku_code": sku_code,
            "size": size,
            "count": count,
            "total_count": total_count,
        }
    )


@login_required
@require_POST
def processing_marking_import(request, order_id: str):
    ok, response = _require_processing_role(request)
    if not ok:
        return response
    latest, payload, agency = _get_processing_order(order_id)
    if not latest:
        return HttpResponseBadRequest("Заявка не найдена")
    file = request.FILES.get("file")
    if not file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return JsonResponse({"ok": False, "error": "Не удалось прочитать .xlsx файл."}, status=400)
    sheet = workbook.active

    rows = []
    barcodes = set()
    invalid_rows = 0
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        barcode = _normalize_cell(row[0]) if row and len(row) > 0 else ""
        code = _normalize_cell(row[1]) if row and len(row) > 1 else ""
        if idx == 1:
            header = barcode.lower()
            if "штрих" in header or "barcode" in header:
                continue
        if not barcode or not code:
            if barcode or code:
                invalid_rows += 1
            continue
        rows.append((barcode, code))
        barcodes.add(barcode)

    if not rows:
        return JsonResponse(
            {"ok": False, "error": "В файле нет данных для импорта."},
            status=400,
        )

    barcode_qs = SKUBarcode.objects.select_related("sku").filter(value__in=barcodes)
    barcode_map = {item.value: item for item in barcode_qs}
    existing_codes = set(
        MarkingCode.objects.filter(code__in=[code for _, code in rows]).values_list("code", flat=True)
    )

    items = extract_processing_items(payload)
    allowed_pairs = {(item["sku_code"], item["size"]) for item in items}
    added = 0
    duplicates = 0
    unknown_barcodes = 0
    mismatched_barcodes = 0
    seen_codes = set()
    to_create = []

    for barcode, code in rows:
        if code in seen_codes:
            duplicates += 1
            continue
        seen_codes.add(code)
        if code in existing_codes:
            duplicates += 1
            continue
        barcode_obj = barcode_map.get(barcode)
        if not barcode_obj or not barcode_obj.sku:
            unknown_barcodes += 1
            continue
        sku_obj = barcode_obj.sku
        if agency and sku_obj.agency and sku_obj.agency_id != agency.id:
            mismatched_barcodes += 1
            continue
        sku_code = sku_obj.sku_code
        size = (barcode_obj.size or sku_obj.size or "").strip()
        if (sku_code, size) not in allowed_pairs:
            if (sku_code, "") in allowed_pairs:
                size = ""
            else:
                matched = [pair for pair in allowed_pairs if pair[0] == sku_code]
                if len(matched) == 1:
                    size = matched[0][1]
                else:
                    unknown_barcodes += 1
                    continue
        to_create.append(
            MarkingCode(
                order_type="processing",
                order_id=order_id,
                agency=agency,
                sku=sku_obj,
                sku_code=sku_code,
                size=size,
                barcode=barcode,
                code=code,
                source="import",
                created_by=request.user if request.user.is_authenticated else None,
            )
        )

    if to_create:
        with transaction.atomic():
            MarkingCode.objects.bulk_create(to_create, batch_size=500)
        added = len(to_create)

    return JsonResponse(
        {
            "ok": True,
            "added": added,
            "duplicates": duplicates,
            "unknown_barcodes": unknown_barcodes,
            "mismatched_barcodes": mismatched_barcodes,
            "invalid_rows": invalid_rows,
        }
    )
