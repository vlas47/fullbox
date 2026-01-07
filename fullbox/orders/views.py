import json
import re
from datetime import datetime, time, timedelta
from pathlib import Path
from urllib.parse import urlencode

from django.conf import settings
from django.db import transaction
from django.http import FileResponse, Http404, HttpResponseForbidden
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.generic import TemplateView
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from audit.models import OrderAuditEntry, log_order_action
from employees.models import Employee
from employees.access import RoleRequiredMixin, get_request_role, resolve_cabinet_url
from sku.models import Agency, SKU, SKUBarcode
from todo.models import Task


_IP_PREFIX_RE = re.compile(r"\bиндивидуальный предприниматель\b", re.IGNORECASE)
_TEMPLATE_DOCS_DIR = settings.BASE_DIR / "static" / "docs"
_ACT_DOCS_DIR = settings.MEDIA_ROOT / "acts"

_ACT_TEMPLATE_FILE = _TEMPLATE_DOCS_DIR / "receiving_act_template.xlsx"
_MX1_TEMPLATE_FILE = _TEMPLATE_DOCS_DIR / "mx1_template.xlsx"
_EXECUTOR_LABEL = (
    'ООО "Фуллбокс", ИНН 5001149130, КПП 503101001, 142450, Московская обл, '
    "г.о. Богородский, г Старая Купавна, ул Магистральная, д. 59, помещ. 55, "
    "тел.: +79778088386"
)
_OKUD_MX1 = "0335001"


def _download_template_file(file_name: str, download_name: str):
    file_path = _TEMPLATE_DOCS_DIR / file_name
    if not file_path.exists():
        raise Http404("Файл не найден")
    return FileResponse(open(file_path, "rb"), as_attachment=True, filename=download_name)


def download_receiving_template(request):
    return _download_template_file("receiving_template.xlsx", "Шаблон приемки.xlsx")


def download_sku_upload_template(request):
    return _download_template_file("sku_upload_template.xlsx", "Шаблон для загрузки номенклатуры.xlsx")


def _shorten_ip_name(name: str) -> str:
    if not name:
        return "-"
    normalized = _IP_PREFIX_RE.sub("ИП", name)
    return " ".join(normalized.split())


def _short_name(full_name: str) -> str:
    if not full_name:
        return "-"
    parts = [part for part in full_name.split() if part]
    if not parts:
        return "-"
    surname = parts[0]
    initials = "".join(f"{part[0].upper()}." for part in parts[1:3] if part)
    return f"{surname} {initials}".strip()


def _order_type_label(order_type: str) -> str:
    if not order_type:
        return "-"
    labels = {"receiving": "ЗП", "packing": "ЗУ"}
    return labels.get(order_type, order_type)


def _first_active_employee_by_roles(*roles: str) -> Employee | None:
    for role in [role for role in roles if role]:
        employee = (
            Employee.objects.filter(role=role, is_active=True)
            .order_by("full_name")
            .first()
        )
        if employee:
            return employee
    return None


def _facsimile_url(employee: Employee | None) -> str:
    if not employee or not getattr(employee, "facsimile", None):
        return ""
    url = employee.facsimile.url
    if url and not url.startswith(("http://", "https://", "/")):
        return f"/{url}"
    return url


def _barcode_value_for_sku(sku, size: str | None) -> str:
    if not sku:
        return "-"
    barcodes = list(getattr(sku, "barcodes", []).all())
    if not barcodes:
        return "-"
    size_value = (size or "").strip()
    if size_value:
        for barcode in barcodes:
            if (barcode.size or "").strip() == size_value:
                return barcode.value
    primary = next((barcode for barcode in barcodes if barcode.is_primary), None)
    return primary.value if primary else barcodes[0].value


def _has_receiving_items(payload: dict) -> bool:
    items = payload.get("items") or []
    for item in items:
        for key in ("sku_code", "name", "qty", "size"):
            if str(item.get(key) or "").strip():
                return True
    return False


def _is_sent_to_manager(payload: dict) -> bool:
    status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
    status_label = (payload.get("status_label") or "").lower()
    if status_value in {"sent_unconfirmed", "send", "submitted"}:
        return True
    return "подтверждени" in status_label


def _order_title_label(order_type: str, payload: dict | None = None) -> str:
    if order_type == "receiving":
        title = "Заявка на приемку"
        if payload and not _has_receiving_items(payload):
            title = "Заявка на приемку без указания товара"
        return title
    if order_type == "packing":
        return "Заявка на упаковку"
    return "Заявка"


def _journal_status_label(entry):
    payload = entry.payload or {}
    status = payload.get("status") or payload.get("submit_action")
    if status == "draft":
        return "Черновик"
    if status in {"sent_unconfirmed", "send", "submitted"}:
        return "Ждет подтверждения"
    return _status_label_from_entry(entry)


def _is_draft_entry(entry) -> bool:
    payload = entry.payload or {}
    status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
    status_label = (payload.get("status_label") or "").lower()
    return status_value == "draft" or "черновик" in status_label


def _can_client_edit_draft(entries, client_agency) -> bool:
    if not client_agency or not entries:
        return False
    if not any(entry.agency_id == client_agency.id for entry in entries if entry.agency_id):
        return False
    status_entry = _current_status_entry(entries)
    return _is_draft_entry(status_entry or entries[-1])


def _is_status_entry(entry) -> bool:
    payload = entry.payload or {}
    if entry.action == "status":
        return True
    return bool(
        payload.get("status")
        or payload.get("status_label")
        or payload.get("submit_action")
    )


def _manager_label() -> str:
    manager = (
        Employee.objects.filter(role="manager", is_active=True)
        .order_by("full_name")
        .first()
    )
    return _short_name(manager.full_name) if manager else "-"


def _storekeeper_label() -> str:
    storekeeper = (
        Employee.objects.filter(role="storekeeper", is_active=True)
        .order_by("full_name")
        .first()
    )
    return _short_name(storekeeper.full_name) if storekeeper else "-"


def _manager_full_name() -> str:
    manager = (
        Employee.objects.filter(role="manager", is_active=True)
        .order_by("full_name")
        .first()
    )
    return manager.full_name.strip() if manager and manager.full_name else ""


def _journal_action_label(entry, manager_label: str, storekeeper_label: str) -> str:
    payload = entry.payload or {}
    status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
    status_label = (payload.get("status_label") or "").lower()
    if status_value == "draft" or "черновик" in status_label:
        return "У клиента"
    if status_value in {"warehouse", "storekeeper"} or "склад" in status_label or "ожидании поставки" in status_label:
        if storekeeper_label and storekeeper_label != "-":
            return f"У кладовщика {storekeeper_label}"
        return "У кладовщика"
    if manager_label and manager_label != "-":
        return f"В работе у менеджера {manager_label}"
    return "В работе у менеджера"


def _status_label_from_entry(entry):
    payload = entry.payload or {}
    client_response = (payload.get("act_client_response") or "").lower()
    if client_response == "confirmed":
        return "Акт приемки подтвержден клиентом"
    if client_response == "dispute":
        return "Клиент заявил разногласия по акту приемки"
    if payload.get("act_sent"):
        return "Акт отправлен клиенту" if not payload.get("act_viewed") else "Выполнена"
    if payload.get("act_storekeeper_signed") and not payload.get("act_manager_signed"):
        return "Принято складом, акт приемки отправлен менеджеру"
    if payload.get("act") == "placement":
        state = (payload.get("act_state") or "closed").lower()
        return "Размещение на складе" if state == "open" else "Товар принят и размещен на складе"
    status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
    status_label = (payload.get("status_label") or "").lower()
    if status_value in {"sent_unconfirmed", "send", "submitted"} or "подтверждени" in status_label:
        return "Ждет подтверждения"
    if status_value in {"warehouse", "on_warehouse"} or "ожидании поставки" in status_label or "на складе" in status_label:
        return "В ожидании поставки товара"
    return payload.get("status_label") or payload.get("status") or "-"


def _current_status_entry(entries):
    for entry in reversed(entries):
        if _is_status_entry(entry):
            return entry
    return entries[-1] if entries else None


def _current_responsible_label(entry):
    if not entry:
        return "-"
    payload = entry.payload or {}
    status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
    status_label = (payload.get("status_label") or "").lower()
    agency = entry.agency
    if status_value == "draft" or "черновик" in status_label:
        base = (agency.fio_agn or agency.agn_name) if agency else None
        base = _shorten_ip_name(base) if base else None
        return f"Клиент {base}" if base else "Клиент"
    if status_value in {"done", "completed", "closed", "finished"} or "выполн" in status_label:
        return "Выполнено"
    if status_value in {"warehouse", "on_warehouse"} or any(token in status_label for token in ("склад", "прием", "приём", "ожидании поставки")):
        return "Склад"
    manager = _manager_full_name()
    return f"Менеджер {manager}" if manager else "Менеджер"


def _actor_label(user=None, agency=None, client_view=False):
    if user:
        base = user.get_full_name() or user.get_username() or str(user)
        base = _shorten_ip_name(base)
        if getattr(user, "is_staff", False):
            return base
        return f"Клиент {base}"
    if client_view and agency:
        base = agency.fio_agn or agency.agn_name or "клиент"
        base = _shorten_ip_name(base)
        return f"Клиент {base}"
    if agency:
        base = agency.fio_agn or agency.agn_name or "клиент"
        base = _shorten_ip_name(base)
        return f"Клиент {base}"
    return "-"


def _history_actor_label(entry, client_view=False, client_label: str | None = None):
    if not entry:
        return "-"
    if entry.action == "create":
        if entry.agency:
            base = entry.agency.fio_agn or entry.agency.agn_name or "клиент"
            base = _shorten_ip_name(base)
            return f"Клиент {base}"
        if client_label and client_label != "-":
            return f"Клиент {client_label}"
        if client_view:
            return "Клиент"
    if entry.action == "update":
        if entry.user:
            base = entry.user.get_full_name() or entry.user.get_username() or str(entry.user)
            base = base.strip()
            if base:
                lowered = base.lower()
                if lowered in {"manager", "менеджер"}:
                    manager_full = _manager_full_name()
                    if manager_full:
                        return f"Менеджер {manager_full}"
                if len(base.split()) >= 2:
                    return f"Менеджер {base}"
                short = _short_name(base)
                if short and short.lower() not in {"manager", "менеджер"}:
                    return f"Менеджер {short}"
        manager_full = _manager_full_name()
        return f"Менеджер {manager_full}" if manager_full else "Менеджер"
    return _actor_label(entry.user, entry.agency, client_view=client_view)


def _history_action_label(entry):
    payload = entry.payload or {}
    act = (payload.get("act") or "").lower()
    if entry.action == "status" and act == "receiving":
        return payload.get("act_label") or "Акт приемки"
    if entry.action == "status" and act == "placement":
        return payload.get("act_label") or "Акт размещения"
    return entry.get_action_display()


def _parse_qty_value(raw: str | None) -> int | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    try:
        return int(text)
    except (TypeError, ValueError):
        return None


def _item_key(sku: str | None, name: str | None, size: str | None) -> str:
    sku_part = (sku or "").strip().lower()
    name_part = (name or "").strip().lower()
    size_part = (size or "").strip().lower()
    return f"{sku_part}|{name_part}|{size_part}"


def _warehouse_status_from_entry(entry) -> bool:
    if not entry:
        return False
    payload = entry.payload or {}
    status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
    status_label = (payload.get("status_label") or "").lower()
    return status_value in {"warehouse", "on_warehouse"} or "склад" in status_label or "ожидании поставки" in status_label


def _act_entry_from_entries(entries, act_type: str = "receiving"):
    for entry in reversed(entries or []):
        if (entry.payload or {}).get("act") == act_type:
            return entry
    return None


def _find_act_entry(entries, act_type: str, label_hint: str):
    entry = _act_entry_from_entries(entries, act_type)
    if entry:
        return entry
    for candidate in reversed(entries or []):
        label = ((candidate.payload or {}).get("act_label") or "").lower()
        if label_hint in label:
            return candidate
    return None


def _is_done_status(entry) -> bool:
    if not entry:
        return False
    payload = entry.payload or {}
    status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
    status_label = (payload.get("status_label") or "").lower()
    if status_value in {"done", "completed", "closed", "finished"}:
        return True
    return "выполн" in status_label


def _act_storekeeper_signed_from_payload(payload: dict) -> bool:
    return bool((payload or {}).get("act_storekeeper_signed"))


def _act_manager_signed_from_payload(payload: dict) -> bool:
    return bool((payload or {}).get("act_manager_signed"))


def _act_storekeeper_signed(entries) -> bool:
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    if not act_entry:
        return False
    return _act_storekeeper_signed_from_payload(act_entry.payload or {})


def _act_manager_signed(entries) -> bool:
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    if not act_entry:
        return False
    return _act_manager_signed_from_payload(act_entry.payload or {})


def _signed_employee_from_payload(payload: dict, employee_key: str) -> Employee | None:
    raw_id = (payload or {}).get(employee_key)
    if raw_id in (None, ""):
        return None
    try:
        employee_id = int(raw_id)
    except (TypeError, ValueError):
        return None
    return Employee.objects.filter(id=employee_id, is_active=True).first()


def _send_act_to_client(order_id, entries, user) -> bool:
    if not entries:
        return False
    receiving_entry = _find_act_entry(entries, "receiving", "акт приемки")
    placement_entry = _find_act_entry(entries, "placement", "акт размещения")
    if not receiving_entry or not placement_entry:
        return False
    if not _placement_closed(entries):
        return False
    receiving_payload = receiving_entry.payload or {}
    if not _act_storekeeper_signed_from_payload(receiving_payload):
        return False
    if not _act_manager_signed_from_payload(receiving_payload):
        return False
    status_entry = _current_status_entry(entries)
    if _is_done_status(status_entry):
        return False
    latest = entries[-1]
    agency = next((entry.agency for entry in reversed(entries) if entry.agency), None)
    if not agency:
        agency = latest.agency
    payload = dict(status_entry.payload or {}) if status_entry else {}
    payload["status"] = "done"
    payload["status_label"] = "Выполнена"
    act_label = (receiving_entry.payload or {}).get("act_label") or "Акт приемки"
    payload["act_sent"] = act_label
    payload["act_sent_at"] = timezone.localtime().isoformat()
    if "act_viewed" not in payload:
        payload["act_viewed"] = False
    log_order_action(
        "status",
        order_id=order_id,
        order_type="receiving",
        user=user if getattr(user, "is_authenticated", False) else None,
        agency=agency,
        description="Акт отправлен клиенту",
        payload=payload,
    )
    log_order_action(
        "update",
        order_id=order_id,
        order_type="receiving",
        user=user if getattr(user, "is_authenticated", False) else None,
        agency=agency,
        description=f"{act_label} отправлен клиенту",
        payload={"message": act_label},
    )
    Task.objects.filter(
        route=f"/orders/receiving/{order_id}/",
        assigned_to__role="manager",
    ).exclude(status="done").update(status="done")
    return True


def _manager_due_date(submitted_at):
    cutoff = submitted_at.replace(hour=14, minute=0, second=0, microsecond=0)
    if submitted_at <= cutoff:
        return submitted_at.replace(hour=18, minute=0, second=0, microsecond=0)
    next_day = submitted_at + timedelta(days=1)
    return next_day.replace(hour=13, minute=0, second=0, microsecond=0)


def _create_manager_task(order_id, agency, request, submitted_at):
    if not agency:
        return
    manager = (
        Employee.objects.filter(role="manager", is_active=True)
        .order_by("full_name")
        .first()
    )
    if not manager:
        return
    description = f"Клиент: {agency.agn_name or agency.inn or agency.id}"
    Task.objects.create(
        title=f"Подтвердите заявку на приемку товара №{order_id}",
        description=description,
        route=f"/orders/receiving/{order_id}/",
        assigned_to=manager,
        created_by=request.user if request.user.is_authenticated else None,
        due_date=_manager_due_date(submitted_at),
    )


def _create_storekeeper_task(order_id, agency, request, submitted_at, observer=None):
    if not agency:
        return
    storekeeper = (
        Employee.objects.filter(role="storekeeper", is_active=True)
        .order_by("full_name")
        .first()
    )
    if not storekeeper:
        return
    description = f"Клиент: {agency.agn_name or agency.inn or agency.id}"
    Task.objects.create(
        title=f"Принять заявку на приемку товара №{order_id}",
        description=description,
        route=f"/orders/receiving/{order_id}/",
        assigned_to=storekeeper,
        observer=observer,
        created_by=request.user if request.user.is_authenticated else None,
        due_date=submitted_at + timedelta(days=1),
    )


def _create_manager_followup_task(order_id, agency, request, submitted_at, observer=None):
    if not agency:
        return
    manager = (
        Employee.objects.filter(role="manager", is_active=True)
        .order_by("full_name")
        .first()
    )
    if not manager:
        return
    title = f"Проверьте размещение по заявке на приемку товара №{order_id}"
    existing = Task.objects.filter(
        route=f"/orders/receiving/{order_id}/",
        assigned_to=manager,
        title=title,
    ).exclude(status="done")
    if existing.exists():
        return
    description = f"Клиент: {agency.agn_name or agency.inn or agency.id}"
    Task.objects.create(
        title=title,
        description=description,
        route=f"/orders/receiving/{order_id}/",
        assigned_to=manager,
        observer=observer,
        created_by=request.user if request.user.is_authenticated else None,
        due_date=_manager_due_date(submitted_at),
    )


def _create_manager_sign_task(order_id, agency, request, observer=None):
    manager = _first_active_employee_by_roles("manager", "head_manager")
    if not manager:
        return
    route = f"/orders/receiving/{order_id}/act/print/"
    description = f"Клиент: {agency.agn_name or agency.inn or agency.id}" if agency else "Клиент: -"
    due_date = timezone.localtime()
    existing = (
        Task.objects.filter(route=route, assigned_to=manager)
        .exclude(status="done")
        .first()
    )
    if existing:
        existing.description = description
        existing.observer = observer
        existing.due_date = due_date
        if existing.status == "done":
            existing.status = "in_progress"
        existing.save()
        return
    Task.objects.create(
        title=f"Подписать акт приемки по заявке №{order_id}",
        description=description,
        route=route,
        assigned_to=manager,
        observer=observer,
        created_by=request.user if request.user.is_authenticated else None,
        due_date=due_date,
        status="in_progress",
    )

def _latest_payload_from_entries(entries):
    for entry in reversed(entries):
        payload = entry.payload or {}
        if not payload:
            continue
        significant_keys = set(payload.keys()) - {
            "comment",
            "message",
            "status",
            "status_label",
            "submit_action",
        }
        if significant_keys:
            return payload
    return entries[-1].payload or {} if entries else {}


def _client_agency_from_request(request):
    if not request.user.is_authenticated:
        return None
    client_id = request.GET.get("client") or request.GET.get("agency")
    if client_id:
        return Agency.objects.filter(pk=client_id, portal_user=request.user).first()
    return Agency.objects.filter(portal_user=request.user).first()


_PHONE_DIGITS_RE = re.compile(r"\D+")


def _is_valid_driver_phone(value: str) -> bool:
    if not value:
        return False
    digits = _PHONE_DIGITS_RE.sub("", value)
    if len(digits) != 11:
        return False
    return digits[0] in {"7", "8"}


def _min_receiving_eta(now):
    tz = timezone.get_current_timezone()
    next_day = (now + timedelta(days=1)).date()
    min_hour = 10
    if now.hour >= 18:
        min_hour = 13
    return timezone.make_aware(datetime.combine(next_day, time(min_hour, 0)), tz)


def _format_payload_value(value):
    if value is None or value == "":
        return "-"
    return str(value).strip()


def _format_party_label(name: str | None, address: str | None, phone: str | None) -> str:
    parts = []
    if name:
        parts.append(name.strip())
    if address:
        parts.append(address.strip())
    if phone:
        parts.append(f"тел.: {phone.strip()}")
    return ", ".join([part for part in parts if part]) or "-"


def _format_payload_list(value):
    if value is None or value == "":
        return "-"
    if isinstance(value, (list, tuple, set)):
        items = [str(item).strip() for item in value if str(item).strip()]
        return ", ".join(items) if items else "-"
    text = str(value).strip()
    return text or "-"


def _resolve_choice_value(value, other):
    other_text = (other or "").strip()
    if other_text:
        return other_text
    return (value or "").strip()


def _place_type_label(value: str) -> str:
    text = _format_payload_value(value)
    if text == "-":
        return text
    labels = {
        "pallet": "Паллет",
        "box": "Короб",
        "bag": "Мешок",
    }
    return labels.get(text, text)


def _format_datetime_value(value):
    text = _format_payload_value(value)
    if text == "-":
        return text
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return text
    return parsed.strftime("%d.%m.%Y, %H:%M")


_ISO_DATETIME_RE = re.compile(r"\b\d{4}-\d{2}-\d{2}T\d{2}:\d{2}(?::\d{2})?\b")


def _format_message_text(text: str) -> str:
    if not text:
        return ""

    def _replace(match):
        raw = match.group(0)
        try:
            parsed = datetime.fromisoformat(raw)
        except ValueError:
            return raw
        return parsed.strftime("%d.%m.%Y, %H:%M")

    return _ISO_DATETIME_RE.sub(_replace, text)


def _normalize_header(value) -> str:
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    return re.sub(r"\s+", " ", text).lower()


def _clean_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    return text


def _header_index(header_map: dict, *names: str):
    for name in names:
        idx = header_map.get(name)
        if idx is not None:
            return idx
    return None


def _load_template_rows(uploaded_file):
    import pandas as pd

    try:
        uploaded_file.seek(0)
    except (AttributeError, OSError):
        pass
    df = pd.read_excel(uploaded_file, header=None, dtype=str)
    if df.empty:
        return {}, []
    header_map = {}
    header_row = df.iloc[0].tolist()
    for idx, value in enumerate(header_row):
        name = _normalize_header(value)
        if not name:
            continue
        header_map.setdefault(name, idx)
    rows = df.iloc[1:].fillna("")
    return header_map, rows


def _template_type_for_header(header_map: dict) -> str | None:
    if "артикул заказчика" in header_map:
        return "sku"
    if "количество" in header_map and (
        "штрихкод" in header_map or "артикул" in header_map or "баркод" in header_map
    ):
        return "receiving"
    return None


def _apply_sku_template(rows, header_map: dict, agency: Agency):
    sku_code_idx = _header_index(header_map, "артикул заказчика", "артикул")
    barcode_idx = _header_index(header_map, "баркод", "штрихкод")
    name_idx = _header_index(header_map, "предмет", "наименование", "товар")
    size_idx = _header_index(header_map, "размер")
    brand_idx = _header_index(header_map, "бренд")
    color_idx = _header_index(header_map, "цвет")
    composition_idx = _header_index(header_map, "состав")
    gender_idx = _header_index(header_map, "пол")
    season_idx = _header_index(header_map, "сезон")
    made_in_idx = _header_index(header_map, "страна пр-ва", "страна производства")
    img_idx = _header_index(header_map, "ссылка на товар", "ссылка")

    if sku_code_idx is None:
        return ["В шаблоне номенклатуры не найдена колонка «Артикул Заказчика»."]

    errors = []
    for _, row in rows.iterrows():
        values = [ _clean_cell(value) for value in row.tolist() ]
        if not any(values):
            continue
        sku_code = _clean_cell(row.iloc[sku_code_idx])
        if not sku_code:
            continue
        name = _clean_cell(row.iloc[name_idx]) if name_idx is not None else ""
        size = _clean_cell(row.iloc[size_idx]) if size_idx is not None else ""
        barcode = _clean_cell(row.iloc[barcode_idx]) if barcode_idx is not None else ""
        brand = _clean_cell(row.iloc[brand_idx]) if brand_idx is not None else ""
        color = _clean_cell(row.iloc[color_idx]) if color_idx is not None else ""
        composition = _clean_cell(row.iloc[composition_idx]) if composition_idx is not None else ""
        gender = _clean_cell(row.iloc[gender_idx]) if gender_idx is not None else ""
        season = _clean_cell(row.iloc[season_idx]) if season_idx is not None else ""
        made_in = _clean_cell(row.iloc[made_in_idx]) if made_in_idx is not None else ""
        img = _clean_cell(row.iloc[img_idx]) if img_idx is not None else ""

        sku = SKU.objects.filter(agency=agency, sku_code=sku_code, deleted=False).first()
        if not sku:
            sku = SKU.objects.create(
                agency=agency,
                sku_code=sku_code,
                name=name or sku_code,
                brand=brand or None,
                size=size or None,
                color=color or None,
                composition=composition or None,
                gender=gender or None,
                season=season or None,
                made_in=made_in or None,
                img=img or None,
                source="manual",
            )
        else:
            updated = False
            if name and not sku.name:
                sku.name = name
                updated = True
            if brand and not sku.brand:
                sku.brand = brand
                updated = True
            if size and not sku.size:
                sku.size = size
                updated = True
            if color and not sku.color:
                sku.color = color
                updated = True
            if composition and not sku.composition:
                sku.composition = composition
                updated = True
            if gender and not sku.gender:
                sku.gender = gender
                updated = True
            if season and not sku.season:
                sku.season = season
                updated = True
            if made_in and not sku.made_in:
                sku.made_in = made_in
                updated = True
            if img and not sku.img:
                sku.img = img
                updated = True
            if updated:
                sku.save(update_fields=[
                    "name",
                    "brand",
                    "size",
                    "color",
                    "composition",
                    "gender",
                    "season",
                    "made_in",
                    "img",
                    "updated_at",
                ])

        if barcode:
            existing_barcode = SKUBarcode.objects.select_related("sku").filter(value=barcode).first()
            if existing_barcode and existing_barcode.sku_id != sku.id:
                errors.append(f"ШК {barcode} уже привязан к SKU {existing_barcode.sku.sku_code}.")
                continue
            if not existing_barcode:
                SKUBarcode.objects.create(
                    sku=sku,
                    value=barcode,
                    size=size or None,
                    is_primary=not sku.barcodes.exists(),
                )

    return errors


def _parse_receiving_template(rows, header_map: dict, agency: Agency):
    barcode_idx = _header_index(header_map, "штрихкод", "баркод")
    qty_idx = _header_index(header_map, "количество", "кол-во", "колво")
    sku_code_idx = _header_index(header_map, "артикул", "артикул заказчика")

    if qty_idx is None or (barcode_idx is None and sku_code_idx is None):
        return [], ["В шаблоне приемки не найдены колонки «Штрихкод» и «Количество»."]

    sku_map = {
        sku.sku_code: sku
        for sku in SKU.objects.filter(agency=agency, deleted=False).prefetch_related("barcodes")
    }
    barcode_map = {
        barcode.value: barcode
        for barcode in SKUBarcode.objects.select_related("sku").filter(sku__agency=agency, sku__deleted=False)
    }

    items = []
    missing = []
    for _, row in rows.iterrows():
        values = [_clean_cell(value) for value in row.tolist()]
        if not any(values):
            continue
        sku_code = _clean_cell(row.iloc[sku_code_idx]) if sku_code_idx is not None else ""
        barcode = _clean_cell(row.iloc[barcode_idx]) if barcode_idx is not None else ""
        qty_raw = _clean_cell(row.iloc[qty_idx]) if qty_idx is not None else ""
        qty_value = _parse_qty_value(qty_raw)
        if not sku_code and not barcode and qty_value is None:
            continue
        if qty_value is None or qty_value <= 0:
            continue

        sku = None
        size_value = ""
        if sku_code:
            sku = sku_map.get(sku_code)
        if not sku and barcode:
            barcode_entry = barcode_map.get(barcode)
            if barcode_entry and barcode_entry.sku and not barcode_entry.sku.deleted:
                sku = barcode_entry.sku
                size_value = (barcode_entry.size or "").strip()
        if not sku:
            missing.append(sku_code or barcode)
            continue
        if not size_value:
            size_value = (sku.size or "").strip()

        items.append(
            {
                "sku_id": str(sku.id),
                "sku_code": sku.sku_code,
                "name": sku.name,
                "size": size_value,
                "qty": str(qty_value),
                "comment": "",
            }
        )

    if missing:
        sample = ", ".join(missing[:5])
        suffix = "..." if len(missing) > 5 else ""
        return [], [f"В шаблоне приемки нет позиций в номенклатуре: {sample}{suffix}"]

    return items, []


def _merge_items(items: list[dict]) -> list[dict]:
    merged = {}
    for item in items:
        sku_code = (item.get("sku_code") or "").strip()
        size = (item.get("size") or "").strip()
        if sku_code:
            key = f"{sku_code.lower()}|{size.lower()}"
        else:
            key = _item_key(item.get("sku_code"), item.get("name"), item.get("size"))
        qty_value = _parse_qty_value(item.get("qty"))
        if key not in merged:
            merged[key] = dict(item)
            if qty_value is not None:
                merged[key]["qty"] = str(qty_value)
            continue
        existing_qty = _parse_qty_value(merged[key].get("qty")) or 0
        merged_qty = existing_qty + (qty_value or 0)
        merged[key]["qty"] = str(merged_qty)
        if not merged[key].get("sku_id") and item.get("sku_id"):
            merged[key]["sku_id"] = item.get("sku_id")
    return list(merged.values())


def _process_template_uploads(template_files, agency: Agency):
    allowed_ext = {".xlsx", ".xls"}
    sku_file = None
    receiving_file = None

    for uploaded_file in template_files:
        if not uploaded_file or not getattr(uploaded_file, "name", ""):
            continue
        ext = "." + uploaded_file.name.split(".")[-1].lower()
        if ext not in allowed_ext:
            raise ValueError(f"Файл «{uploaded_file.name}» не является Excel-шаблоном.")
        header_map, rows = _load_template_rows(uploaded_file)
        template_type = _template_type_for_header(header_map)
        if not template_type:
            raise ValueError(f"Файл «{uploaded_file.name}» не похож на шаблон приемки или номенклатуры.")
        if template_type == "sku":
            if sku_file:
                raise ValueError("Можно загрузить только один шаблон номенклатуры.")
            sku_file = (rows, header_map, uploaded_file.name)
        elif template_type == "receiving":
            if receiving_file:
                raise ValueError("Можно загрузить только один шаблон приемки.")
            receiving_file = (rows, header_map, uploaded_file.name)

    template_names = []
    for entry in (sku_file, receiving_file):
        if entry:
            template_names.append(entry[2])

    if sku_file:
        rows, header_map, _ = sku_file
        errors = _apply_sku_template(rows, header_map, agency)
        if errors:
            raise ValueError("; ".join(errors))

    items = []
    if receiving_file:
        rows, header_map, _ = receiving_file
        items, errors = _parse_receiving_template(rows, header_map, agency)
        if errors:
            raise ValueError("; ".join(errors))

    return items, template_names


def _describe_payload_changes(old_payload, new_payload):
    changes = []
    fields = [
        ("eta_at", "Плановая дата/время"),
        ("expected_boxes", "Количество мест"),
        ("place_type", "Тип мест"),
        ("vehicle_number", "Номер авто"),
        ("driver_phone", "Телефон водителя"),
        ("comment", "Комментарий"),
    ]
    for key, label in fields:
        if key == "eta_at":
            old_val = _format_datetime_value((old_payload or {}).get(key))
            new_val = _format_datetime_value((new_payload or {}).get(key))
        elif key == "place_type":
            old_val = _place_type_label((old_payload or {}).get(key))
            new_val = _place_type_label((new_payload or {}).get(key))
        else:
            old_val = _format_payload_value((old_payload or {}).get(key))
            new_val = _format_payload_value((new_payload or {}).get(key))
        if old_val != new_val:
            changes.append(f"{label}: {old_val} → {new_val}")
    old_items = (old_payload or {}).get("items") or []
    new_items = (new_payload or {}).get("items") or []
    if old_items != new_items:
        changes.append(f"Состав поставки: {len(old_items)} → {len(new_items)} поз.")
    return changes


def _safe_doc_name(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value or "")).strip("_")
    return text or "act"


def _format_doc_date(value: str | None) -> str:
    if not value:
        return timezone.localtime().strftime("%d.%m.%Y")
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return str(value)
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return timezone.localtime(parsed).strftime("%d.%m.%Y")


def _agency_label(agency: Agency | None) -> str:
    if not agency:
        return "-"
    parts = []
    name = (agency.agn_name or agency.fio_agn or "").strip()
    if name:
        parts.append(name)
    inn = (agency.inn or "").strip()
    if inn:
        parts.append(f"ИНН {inn}")
    return ", ".join(parts) if parts else f"Клиент {agency.id}"


def _placement_box_map(placement_payload: dict | None) -> dict:
    items = (placement_payload or {}).get("act_items") or []
    box_map = {}
    for item in items:
        key = _item_key(item.get("sku_code"), item.get("name"), item.get("size"))
        if not key:
            continue
        box_map[key] = _parse_qty_value(item.get("box_qty")) or 0
    return box_map


def _placement_box_count_map(placement_payload: dict | None) -> dict:
    boxes = (placement_payload or {}).get("act_boxes") or []
    box_counts: dict[str, int] = {}
    for box in boxes:
        if not isinstance(box, dict):
            continue
        keys = set()
        for item in (box.get("items") or []):
            if not isinstance(item, dict):
                continue
            sku_code = item.get("sku") or item.get("sku_code")
            key = _item_key(sku_code, item.get("name"), item.get("size"))
            if key:
                keys.add(key)
        for key in keys:
            box_counts[key] = box_counts.get(key, 0) + 1
    return box_counts


def _placement_pallet_count_map(placement_payload: dict | None) -> dict:
    pallets = (placement_payload or {}).get("act_pallets") or []
    pallet_counts: dict[str, int] = {}
    for pallet in pallets:
        if not isinstance(pallet, dict):
            continue
        keys = set()
        for item in (pallet.get("items") or []):
            if not isinstance(item, dict):
                continue
            sku_code = item.get("sku") or item.get("sku_code")
            key = _item_key(sku_code, item.get("name"), item.get("size"))
            if key:
                keys.add(key)
        for key in keys:
            pallet_counts[key] = pallet_counts.get(key, 0) + 1
    return pallet_counts


def _act_items_with_barcodes(
    agency_id: int | None,
    act_items: list[dict],
    placement_box_map: dict | None = None,
) -> list[dict]:
    sku_codes = set()
    for item in act_items:
        sku_code = str(item.get("sku_code") or "").strip()
        if sku_code:
            sku_codes.add(sku_code)

    sku_by_code = {}
    if agency_id and sku_codes:
        sku_qs = SKU.objects.filter(agency_id=agency_id, sku_code__in=sku_codes, deleted=False).prefetch_related("barcodes")
        for sku in sku_qs:
            sku_by_code[sku.sku_code] = sku

    rows = []
    for item in act_items:
        sku_code = str(item.get("sku_code") or "").strip()
        size = str(item.get("size") or "").strip()
        name = str(item.get("name") or "").strip()
        barcode = str(item.get("barcode") or "").strip()
        if not barcode and sku_code:
            sku = sku_by_code.get(sku_code)
            barcode = _barcode_value_for_sku(sku, size)
        planned = _parse_qty_value(item.get("planned_qty") or item.get("qty")) or 0
        actual = _parse_qty_value(item.get("actual_qty") or item.get("qty")) or 0
        key = _item_key(sku_code, name, size)
        box_qty = None
        if placement_box_map and key in placement_box_map:
            box_qty = placement_box_map.get(key)
        rows.append(
            {
                "sku_code": sku_code,
                "name": name,
                "size": size,
                "barcode": barcode,
                "planned": planned,
                "actual": actual,
                "box_qty": box_qty,
            }
        )
    return rows


def _merge_row_cells(ws, row: int, left: str, right: str):
    cell_range = f"{left}{row}:{right}{row}"
    if cell_range in ws.merged_cells:
        return
    ws.merge_cells(cell_range)


def _set_cell_value(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _find_row_with_text(ws, needles: tuple[str, ...], start_row: int) -> int | None:
    lowered = [needle.lower() for needle in needles]
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            text = cell.value.strip().lower()
            if not text:
                continue
            if any(needle in text for needle in lowered):
                return cell.row
    return None


def _fill_receiving_act_sheet(
    ws,
    order_id: str,
    agency: Agency | None,
    act_items: list[dict],
    eta_raw: str | None,
    expected_boxes: int | None = None,
):
    doc_date = _format_doc_date(eta_raw)
    ws["B7"] = f"Приемка № {order_id} от {doc_date}"
    ws["B9"] = f"Поставщик: {_agency_label(agency)}"

    start_row = 14
    footer_row = _find_row_with_text(ws, ("итого",), start_row) or ws.max_row + 1
    current_rows = max(0, footer_row - start_row)
    item_count = len(act_items)
    if item_count < current_rows:
        ws.delete_rows(start_row + item_count, current_rows - item_count)
    elif item_count > current_rows:
        ws.insert_rows(footer_row, item_count - current_rows)
    footer_row = start_row + item_count
    for row in range(start_row, footer_row):
        for col in range(2, 13):
            _set_cell_value(ws, row, col, None)
    for idx, item in enumerate(act_items):
        row = start_row + idx
        _merge_row_cells(ws, row, "C", "D")
        _merge_row_cells(ws, row, "E", "I")
        ws[f"B{row}"] = idx + 1
        ws[f"C{row}"] = item.get("barcode") or "-"
        name_parts = [item.get("name") or "-"]
        if item.get("sku_code"):
            name_parts.append(f"арт.: {item['sku_code']}")
        if item.get("size"):
            name_parts.append(f"р-р {item['size']}")
        ws[f"E{row}"] = ", ".join(name_parts)
        ws[f"J{row}"] = item.get("planned", 0)
        ws[f"K{row}"] = item.get("actual", 0)
        box_qty = item.get("box_qty")
        if box_qty is not None:
            ws[f"L{row}"] = box_qty
    total_planned = sum(item.get("planned", 0) or 0 for item in act_items)
    total_actual = sum(item.get("actual", 0) or 0 for item in act_items)
    total_boxes = sum((item.get("box_qty") or 0) for item in act_items if item.get("box_qty") is not None)
    if not total_boxes and expected_boxes:
        total_boxes = expected_boxes
    ws[f"J{footer_row}"] = total_planned
    ws[f"K{footer_row}"] = total_actual
    if total_boxes:
        ws[f"L{footer_row}"] = total_boxes


def _fill_mx1_sheet_one(ws, order_id: str, agency: Agency | None, act_items: list[dict]):
    ws["H18"] = order_id
    ws["C12"] = _agency_label(agency)

    start_row = 29
    capacity = 20
    items = act_items[:capacity]
    footer_row = _find_row_with_text(ws, ("итого", "èòîãî"), start_row) or (start_row + capacity)
    current_rows = max(0, footer_row - start_row)
    item_count = len(items)
    if item_count < current_rows:
        ws.delete_rows(start_row + item_count, current_rows - item_count)
    elif item_count > current_rows:
        ws.insert_rows(footer_row, item_count - current_rows)
    footer_row = start_row + item_count
    for row in range(start_row, footer_row):
        for col in range(2, 13):
            _set_cell_value(ws, row, col, None)
    for idx, item in enumerate(items):
        row = start_row + idx
        _merge_row_cells(ws, row, "B", "C")
        _merge_row_cells(ws, row, "G", "H")
        _merge_row_cells(ws, row, "I", "K")
        _merge_row_cells(ws, row, "L", "O")
        ws[f"B{row}"] = idx + 1
        name_parts = [item.get("name") or "-"]
        if item.get("sku_code"):
            name_parts.append(f"арт.: {item['sku_code']}")
        ws[f"D{row}"] = ", ".join(name_parts)
        characteristic = item.get("size") or ""
        if characteristic:
            ws[f"F{row}"] = f"размер {characteristic}"
        ws[f"G{row}"] = "шт"
        ws[f"I{row}"] = "796"
        ws[f"L{row}"] = item.get("actual", 0)
    ws[f"L{footer_row}"] = sum(item.get("actual", 0) or 0 for item in items)


def _fill_mx1_sheet_two(ws, act_items: list[dict], offset: int):
    start_row = 7
    capacity = 23
    items = act_items[offset:offset + capacity]
    footer_row = _find_row_with_text(ws, ("условия хранения",), start_row) or (start_row + capacity)
    current_rows = max(0, footer_row - start_row)
    item_count = len(items)
    if item_count < current_rows:
        ws.delete_rows(start_row + item_count, current_rows - item_count)
    elif item_count > current_rows:
        ws.insert_rows(footer_row, item_count - current_rows)
    footer_row = start_row + item_count
    for row in range(start_row, footer_row):
        for col in range(2, 15):
            _set_cell_value(ws, row, col, None)
    for idx, item in enumerate(items):
        row = start_row + idx
        _merge_row_cells(ws, row, "C", "F")
        _merge_row_cells(ws, row, "H", "I")
        _merge_row_cells(ws, row, "J", "L")
        ws[f"B{row}"] = offset + idx + 1
        name_parts = [item.get("name") or "-"]
        if item.get("sku_code"):
            name_parts.append(f"арт.: {item['sku_code']}")
        ws[f"C{row}"] = ", ".join(name_parts)
        characteristic = item.get("size") or ""
        if characteristic:
            ws[f"H{row}"] = f"размер {characteristic}"
        ws[f"J{row}"] = "шт"
        ws[f"M{row}"] = "796"
        ws[f"N{row}"] = item.get("actual", 0)


def _ensure_act_documents(
    order_id: str,
    agency: Agency | None,
    act_payload: dict,
    placement_payload: dict | None = None,
):
    if not _ACT_TEMPLATE_FILE.exists() or not _MX1_TEMPLATE_FILE.exists():
        raise Http404("Шаблоны актов не найдены.")
    placement_box_map = _placement_box_map(placement_payload)
    act_items = _act_items_with_barcodes(
        agency.id if agency else None,
        act_payload.get("act_items") or [],
        placement_box_map,
    )
    expected_boxes = _parse_qty_value(act_payload.get("expected_boxes"))
    safe_id = _safe_doc_name(order_id)
    target_dir = Path(_ACT_DOCS_DIR) / safe_id
    target_dir.mkdir(parents=True, exist_ok=True)
    act_path = target_dir / f"act_receiving_{safe_id}.xlsx"
    mx1_path = target_dir / f"mx1_{safe_id}.xlsx"
    wb = load_workbook(_ACT_TEMPLATE_FILE)
    ws = wb.active
    _fill_receiving_act_sheet(
        ws,
        order_id,
        agency,
        act_items,
        act_payload.get("eta_at"),
        expected_boxes,
    )
    wb.save(act_path)
    wb = load_workbook(_MX1_TEMPLATE_FILE)
    sheet_one = wb["МХ-1 (1стр)"]
    sheet_two = wb["МХ-1(2стр)"]
    _fill_mx1_sheet_one(sheet_one, order_id, agency, act_items)
    _fill_mx1_sheet_two(sheet_two, act_items, 20)
    wb.save(mx1_path)
    return act_path, mx1_path


def _load_order_entries(order_id: str, order_type: str = "receiving"):
    return list(
        OrderAuditEntry.objects.filter(order_id=order_id, order_type=order_type).order_by("created_at")
    )


def _act_access_allowed(request, entries):
    if not request.user.is_authenticated:
        return False
    client_agency = _client_agency_from_request(request)
    if client_agency:
        return any(entry.agency_id == client_agency.id for entry in entries if entry.agency_id)
    role = get_request_role(request)
    return request.user.is_staff or role in {"storekeeper", "manager", "head_manager", "director", "admin"}


def _placement_closed(entries) -> bool:
    placement_entry = _find_act_entry(entries, "placement", "акт размещения")
    if not placement_entry:
        return False
    state = ((placement_entry.payload or {}).get("act_state") or "closed").lower()
    return state == "closed"


def download_receiving_act_doc(request, order_id: str):
    entries = _load_order_entries(order_id)
    if not entries:
        raise Http404("Заявка не найдена")
    if not _act_access_allowed(request, entries):
        return HttpResponseForbidden("Доступ запрещен")
    if not _placement_closed(entries):
        return redirect(f"/orders/receiving/{order_id}/act/?error=placement_required")
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    if not act_entry:
        raise Http404("Акт приемки не найден")
    placement_entry = _find_act_entry(entries, "placement", "акт размещения")
    act_path, _ = _ensure_act_documents(
        order_id,
        act_entry.agency,
        act_entry.payload or {},
        placement_entry.payload if placement_entry else None,
    )
    return FileResponse(open(act_path, "rb"), as_attachment=True, filename=act_path.name)


def download_receiving_act_mx1(request, order_id: str):
    entries = _load_order_entries(order_id)
    if not entries:
        raise Http404("Заявка не найдена")
    if not _act_access_allowed(request, entries):
        return HttpResponseForbidden("Доступ запрещен")
    if not _placement_closed(entries):
        return redirect(f"/orders/receiving/{order_id}/act/?error=placement_required")
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    if not act_entry:
        raise Http404("Акт приемки не найден")
    placement_entry = _find_act_entry(entries, "placement", "акт размещения")
    _, mx1_path = _ensure_act_documents(
        order_id,
        act_entry.agency,
        act_entry.payload or {},
        placement_entry.payload if placement_entry else None,
    )
    return FileResponse(open(mx1_path, "rb"), as_attachment=True, filename=mx1_path.name)


def print_receiving_act(request, order_id: str):
    entries = _load_order_entries(order_id)
    if not entries:
        raise Http404("Заявка не найдена")
    if not _act_access_allowed(request, entries):
        return HttpResponseForbidden("Доступ запрещен")
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    if not act_entry:
        raise Http404("Акт приемки не найден")
    placement_entry = _find_act_entry(entries, "placement", "акт размещения")
    if not _placement_closed(entries):
        return redirect(f"/orders/receiving/{order_id}/act/?error=placement_required")
    act_payload = act_entry.payload or {}
    placement_payload = placement_entry.payload if placement_entry else None
    agency = act_entry.agency or (entries[-1].agency if entries else None)
    doc_date = _format_doc_date(act_payload.get("eta_at") or act_entry.created_at.isoformat())
    arrival_at = _format_datetime_value(act_payload.get("eta_at"))
    vehicle_number = _format_payload_value(act_payload.get("vehicle_number"))
    driver_phone = _format_payload_value(act_payload.get("driver_phone"))
    agency_name = "-"
    agency_inn = "-"
    agency_kpp = "-"
    agency_address = "-"
    if agency:
        agency_name = (agency.agn_name or agency.fio_agn or str(agency)).strip() or agency_name
        agency_inn = (agency.inn or "").strip() or agency_inn
        agency_kpp = (agency.kpp or "").strip() or agency_kpp
        agency_address = (agency.adres or agency.fakt_adres or "").strip() or agency_address

    act_items_raw = act_payload.get("act_items") or []
    box_count_map = _placement_box_count_map(placement_payload) if placement_payload else None
    boxes_payload = (placement_payload or {}).get("act_boxes") or []
    pallets_payload = (placement_payload or {}).get("act_pallets") or []
    box_codes = set()
    for box in boxes_payload:
        if not isinstance(box, dict):
            continue
        code = (box.get("code") or "").strip()
        if code:
            box_codes.add(code)
    if box_codes:
        total_boxes_unique = len(box_codes)
    else:
        total_boxes_unique = len([box for box in boxes_payload if isinstance(box, dict)])
    total_pallets = len([pallet for pallet in pallets_payload if isinstance(pallet, dict)])
    act_items = _act_items_with_barcodes(agency.id if agency else None, act_items_raw, box_count_map)
    display_items = []
    total_planned = 0
    total_actual = 0
    total_boxes = 0
    for item in act_items:
        planned = item.get("planned", 0) or 0
        actual = item.get("actual", 0) or 0
        total_planned += planned
        total_actual += actual
        box_qty = item.get("box_qty")
        if box_qty is not None:
            total_boxes += box_qty
        mismatch = actual - planned
        name_parts = [item.get("name") or "-"]
        if item.get("size"):
            name_parts.append(f"р-р {item['size']}")
        display_items.append(
            {
                "barcode": item.get("barcode") or "-",
                "sku_code": item.get("sku_code") or "-",
                "name": ", ".join(name_parts),
                "planned": planned,
                "actual": actual,
                "box_qty": "-" if box_qty is None else box_qty,
                "mismatch": "" if mismatch == 0 else mismatch,
            }
        )
    total_boxes_label = total_boxes_unique if total_boxes_unique else "-"
    total_pallets_label = total_pallets if total_pallets else "-"
    total_mismatch = total_actual - total_planned

    storekeeper_signed = _act_storekeeper_signed_from_payload(act_payload)
    manager_signed = _act_manager_signed_from_payload(act_payload)
    storekeeper_employee = (
        _signed_employee_from_payload(act_payload, "act_storekeeper_employee_id")
        if storekeeper_signed
        else None
    )
    manager_employee = (
        _signed_employee_from_payload(act_payload, "act_manager_employee_id")
        if manager_signed
        else None
    )
    role = get_request_role(request)
    client_agency = _client_agency_from_request(request)
    client_view = bool(client_agency)
    if storekeeper_signed and not manager_signed:
        _create_manager_sign_task(order_id, agency, request, observer=storekeeper_employee)
    can_storekeeper_sign = role == "storekeeper" and not storekeeper_signed and not client_view
    can_manager_sign = (
        role in {"manager", "head_manager", "director", "admin"}
        and storekeeper_signed
        and not manager_signed
        and not client_view
    )

    client_agency = _client_agency_from_request(request)
    client_view = bool(client_agency)
    default_return_url = f"/orders/receiving/{order_id}/act/"
    if client_agency:
        default_return_url = f"/client/dashboard/?client={client_agency.id}"
    raw_return_url = request.GET.get("return") or request.META.get("HTTP_REFERER")
    return_url = _safe_return_url(request, raw_return_url, request.path) or default_return_url
    status_entry = _current_status_entry(entries)
    status_payload = status_entry.payload or {} if status_entry else {}
    client_response = (
        status_payload.get("act_client_response")
        or request.GET.get("response")
        or ""
    ).lower()
    sent_at_raw = status_payload.get("act_sent_at") or act_entry.created_at.isoformat()
    client_deadline = ""
    try:
        sent_at = datetime.fromisoformat(str(sent_at_raw))
        if timezone.is_naive(sent_at):
            sent_at = timezone.make_aware(sent_at, timezone.get_current_timezone())
        sent_at = timezone.localtime(sent_at)
        client_deadline = (sent_at + timedelta(hours=24)).strftime("%d.%m.%Y %H:%M")
    except (TypeError, ValueError):
        client_deadline = ""

    ctx = {
        "order_id": order_id,
        "act_label": (act_payload.get("act_label") or "Акт приемки").strip(),
        "doc_date": doc_date,
        "arrival_at": arrival_at,
        "vehicle_number": vehicle_number,
        "driver_phone": driver_phone,
        "agency_name": agency_name,
        "agency_inn": agency_inn,
        "agency_kpp": agency_kpp,
        "agency_address": agency_address,
        "agency_label": _agency_label(agency),
        "items": display_items,
        "total_planned": total_planned,
        "total_actual": total_actual,
        "total_boxes": total_boxes_label,
        "total_pallets": total_pallets_label,
        "total_mismatch": "" if total_mismatch == 0 else total_mismatch,
        "print_date": timezone.localtime().strftime("%d.%m.%Y %H:%M"),
        "executor_label": _EXECUTOR_LABEL,
        "manager_facsimile_url": _facsimile_url(manager_employee),
        "storekeeper_facsimile_url": _facsimile_url(storekeeper_employee),
        "storekeeper_signed": storekeeper_signed,
        "manager_signed": manager_signed,
        "can_storekeeper_sign": can_storekeeper_sign,
        "can_manager_sign": can_manager_sign,
        "sign_status": request.GET.get("signed"),
        "sign_error": request.GET.get("error"),
        "return_url": return_url,
        "client_view": client_view,
        "client_param": client_agency.id if client_agency else "",
        "client_response": client_response,
        "client_deadline": client_deadline,
    }
    return render(request, "orders/receiving_act_print.html", ctx)


def sign_receiving_act_storekeeper(request, order_id: str):
    if request.method != "POST":
        return HttpResponseForbidden("Доступ запрещен")
    entries = _load_order_entries(order_id)
    if not entries:
        raise Http404("Заявка не найдена")
    if not _act_access_allowed(request, entries):
        return HttpResponseForbidden("Доступ запрещен")
    role = get_request_role(request)
    if role != "storekeeper":
        return HttpResponseForbidden("Доступ запрещен")
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    if not act_entry:
        raise Http404("Акт приемки не найден")
    if not _placement_closed(entries):
        return redirect(f"/orders/receiving/{order_id}/act/print/?error=placement_required")
    act_payload = dict(act_entry.payload or {})
    if _act_storekeeper_signed_from_payload(act_payload):
        return redirect(f"/orders/receiving/{order_id}/act/print/")
    employee = Employee.objects.filter(user=request.user, role="storekeeper", is_active=True).first()
    if not employee:
        employee = _first_active_employee_by_roles("storekeeper")
    if not employee or not getattr(employee, "facsimile", None):
        return redirect(f"/orders/receiving/{order_id}/act/print/?error=storekeeper_facsimile")
    act_payload["act_storekeeper_signed"] = True
    act_payload["act_storekeeper_signed_at"] = timezone.localtime().isoformat()
    act_payload["act_storekeeper_employee_id"] = employee.id
    if request.user.is_authenticated:
        act_payload["act_storekeeper_user_id"] = request.user.id
    if not act_payload.get("status"):
        act_payload["status"] = "warehouse"
    act_payload["status_label"] = "Принято складом, акт приемки отправлен менеджеру"
    log_order_action(
        "status",
        order_id=order_id,
        order_type="receiving",
        user=request.user if request.user.is_authenticated else None,
        agency=act_entry.agency or (entries[-1].agency if entries else None),
        description="Акт приемки подписан кладовщиком и отправлен менеджеру",
        payload=act_payload,
    )
    _create_manager_sign_task(
        order_id,
        act_entry.agency or (entries[-1].agency if entries else None),
        request,
        observer=employee,
    )
    return redirect(f"/orders/receiving/{order_id}/act/print/?signed=storekeeper")


def sign_receiving_act_manager(request, order_id: str):
    if request.method != "POST":
        return HttpResponseForbidden("Доступ запрещен")
    entries = _load_order_entries(order_id)
    if not entries:
        raise Http404("Заявка не найдена")
    if not _act_access_allowed(request, entries):
        return HttpResponseForbidden("Доступ запрещен")
    role = get_request_role(request)
    if role not in {"manager", "head_manager", "director", "admin"}:
        return HttpResponseForbidden("Доступ запрещен")
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    if not act_entry:
        raise Http404("Акт приемки не найден")
    if not _placement_closed(entries):
        return redirect(f"/orders/receiving/{order_id}/act/print/?error=placement_required")
    act_payload = dict(act_entry.payload or {})
    if not _act_storekeeper_signed_from_payload(act_payload):
        return redirect(f"/orders/receiving/{order_id}/act/print/?error=storekeeper_required")
    if _act_manager_signed_from_payload(act_payload):
        return redirect(f"/orders/receiving/{order_id}/act/print/?signed=manager")
    employee = Employee.objects.filter(user=request.user, is_active=True).first()
    if not employee:
        employee = _first_active_employee_by_roles("manager", "head_manager")
    if not employee or not getattr(employee, "facsimile", None):
        return redirect(f"/orders/receiving/{order_id}/act/print/?error=manager_facsimile")
    act_payload["act_manager_signed"] = True
    act_payload["act_manager_signed_at"] = timezone.localtime().isoformat()
    act_payload["act_manager_employee_id"] = employee.id
    if request.user.is_authenticated:
        act_payload["act_manager_user_id"] = request.user.id
    log_order_action(
        "status",
        order_id=order_id,
        order_type="receiving",
        user=request.user if request.user.is_authenticated else None,
        agency=act_entry.agency or (entries[-1].agency if entries else None),
        description="Акт приемки подписан менеджером",
        payload=act_payload,
    )
    entries = _load_order_entries(order_id)
    sent = _send_act_to_client(order_id, entries, request.user)
    if not sent:
        return redirect(f"/orders/receiving/{order_id}/act/print/?signed=manager&error=send")
    Task.objects.filter(
        route=f"/orders/receiving/{order_id}/act/print/",
        assigned_to__role__in=["manager", "head_manager"],
    ).delete()
    return redirect(f"/orders/receiving/{order_id}/act/print/?signed=manager")


def _client_print_url(
    order_id: str,
    client_agency,
    response: str | None = None,
    return_url: str | None = None,
) -> str:
    params = {}
    if client_agency:
        params["client"] = client_agency.id
    if response:
        params["response"] = response
    if return_url:
        params["return"] = return_url
    suffix = f"?{urlencode(params)}" if params else ""
    return f"/orders/receiving/{order_id}/act/print/{suffix}"


def _safe_return_url(request, raw_url: str | None, current_path: str | None = None) -> str:
    if not raw_url:
        return ""
    if current_path:
        current_abs = request.build_absolute_uri(current_path)
        if raw_url.startswith(current_abs) or raw_url.startswith(current_path):
            return ""
    if url_has_allowed_host_and_scheme(
        raw_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return raw_url
    return ""


def _client_act_access_allowed(request, entries):
    client_agency = _client_agency_from_request(request)
    if not client_agency:
        return False
    if not any(entry.agency_id == client_agency.id for entry in entries if entry.agency_id):
        return False
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    return bool(act_entry)


def confirm_receiving_act_client(request, order_id: str):
    if request.method != "POST":
        return HttpResponseForbidden("Доступ запрещен")
    entries = _load_order_entries(order_id)
    if not entries:
        raise Http404("Заявка не найдена")
    if not _client_act_access_allowed(request, entries):
        return HttpResponseForbidden("Доступ запрещен")
    client_agency = _client_agency_from_request(request)
    raw_return_url = request.POST.get("return") or request.GET.get("return")
    return_url = _safe_return_url(request, raw_return_url, request.path)
    status_entry = _current_status_entry(entries)
    payload = dict(status_entry.payload or {}) if status_entry else {}
    existing_response = (payload.get("act_client_response") or "").lower()
    if existing_response in {"confirmed", "dispute"}:
        return redirect(_client_print_url(order_id, client_agency, existing_response, return_url))
    payload["act_client_response"] = "confirmed"
    payload["act_client_response_at"] = timezone.localtime().isoformat()
    log_order_action(
        "status",
        order_id=order_id,
        order_type="receiving",
        user=request.user if request.user.is_authenticated else None,
        agency=entries[-1].agency if entries else None,
        description="Акт приемки подтвержден клиентом",
        payload=payload,
    )
    return redirect(_client_print_url(order_id, client_agency, "confirmed", return_url))


def dispute_receiving_act_client(request, order_id: str):
    if request.method != "POST":
        return HttpResponseForbidden("Доступ запрещен")
    entries = _load_order_entries(order_id)
    if not entries:
        raise Http404("Заявка не найдена")
    if not _client_act_access_allowed(request, entries):
        return HttpResponseForbidden("Доступ запрещен")
    client_agency = _client_agency_from_request(request)
    raw_return_url = request.POST.get("return") or request.GET.get("return")
    return_url = _safe_return_url(request, raw_return_url, request.path)
    status_entry = _current_status_entry(entries)
    payload = dict(status_entry.payload or {}) if status_entry else {}
    existing_response = (payload.get("act_client_response") or "").lower()
    if existing_response in {"confirmed", "dispute"}:
        return redirect(_client_print_url(order_id, client_agency, existing_response, return_url))
    payload["act_client_response"] = "dispute"
    payload["act_client_response_at"] = timezone.localtime().isoformat()
    log_order_action(
        "status",
        order_id=order_id,
        order_type="receiving",
        user=request.user if request.user.is_authenticated else None,
        agency=entries[-1].agency if entries else None,
        description="Клиент заявил разногласия по акту приемки",
        payload=payload,
    )
    return redirect(_client_print_url(order_id, client_agency, "dispute", return_url))


def print_receiving_act_mx1(request, order_id: str):
    entries = _load_order_entries(order_id)
    if not entries:
        raise Http404("Заявка не найдена")
    if not _act_access_allowed(request, entries):
        return HttpResponseForbidden("Доступ запрещен")
    act_entry = _find_act_entry(entries, "receiving", "акт приемки")
    if not act_entry:
        raise Http404("Акт приемки не найден")
    placement_entry = _find_act_entry(entries, "placement", "акт размещения")
    if not _placement_closed(entries):
        return redirect(f"/orders/receiving/{order_id}/act/?error=placement_required")
    act_payload = act_entry.payload or {}
    placement_payload = placement_entry.payload if placement_entry else None
    agency = act_entry.agency or (entries[-1].agency if entries else None)
    doc_date = _format_doc_date(act_payload.get("eta_at") or act_entry.created_at.isoformat())
    agency_name = "-"
    agency_address = "-"
    agency_phone = "-"
    if agency:
        agency_name = (agency.agn_name or agency.fio_agn or str(agency)).strip() or agency_name
        agency_address = (agency.adres or agency.fakt_adres or "").strip() or agency_address
        agency_phone = (agency.phone or "").strip() or agency_phone
    keeper_label = _EXECUTOR_LABEL
    depositor_label = _format_party_label(agency_name, agency_address, agency_phone)

    act_items_raw = act_payload.get("act_items") or []
    box_count_map = _placement_box_count_map(placement_payload) if placement_payload else {}
    pallet_count_map = _placement_pallet_count_map(placement_payload) if placement_payload else {}
    act_items = _act_items_with_barcodes(agency.id if agency else None, act_items_raw, box_count_map)
    display_items = []
    total_actual = 0
    for item in act_items:
        actual = item.get("actual", 0) or 0
        total_actual += actual
        key = _item_key(item.get("sku_code"), item.get("name"), item.get("size"))
        pallet_qty = pallet_count_map.get(key)
        name_parts = [item.get("name") or "-"]
        if item.get("sku_code"):
            name_parts.append(f"арт.: {item['sku_code']}")
        if item.get("size"):
            name_parts.append(f"р-р {item['size']}")
        display_items.append(
            {
                "sku_code": item.get("sku_code") or "-",
                "barcode": item.get("barcode") or "-",
                "name": ", ".join(name_parts),
                "size": item.get("size") or "-",
                "actual": actual,
                "box_qty": "-" if item.get("box_qty") is None else item.get("box_qty"),
                "pallet_qty": "-" if pallet_qty is None else pallet_qty,
            }
        )

    boxes_payload = (placement_payload or {}).get("act_boxes") or []
    pallets_payload = (placement_payload or {}).get("act_pallets") or []
    box_codes = set()
    for box in boxes_payload:
        if not isinstance(box, dict):
            continue
        code = (box.get("code") or "").strip()
        if code:
            box_codes.add(code)
    if box_codes:
        total_boxes = len(box_codes)
    else:
        total_boxes = len([box for box in boxes_payload if isinstance(box, dict)])
    total_pallets = len([pallet for pallet in pallets_payload if isinstance(pallet, dict)])

    manager_employee = _first_active_employee_by_roles("manager", "head_manager")
    storekeeper_employee = _first_active_employee_by_roles("storekeeper")

    items_per_page = 18
    pages = []
    for start in range(0, max(len(display_items), 1), items_per_page):
        pages.append(
            {
                "start": start,
                "items": display_items[start:start + items_per_page],
            }
        )

    ctx = {
        "order_id": order_id,
        "doc_date": doc_date,
        "agency_label": _agency_label(agency),
        "executor_label": _EXECUTOR_LABEL,
        "okud_code": _OKUD_MX1,
        "keeper_label": keeper_label,
        "depositor_label": depositor_label,
        "pages": pages,
        "total_actual": total_actual,
        "total_boxes": total_boxes if total_boxes else "-",
        "total_pallets": total_pallets if total_pallets else "-",
        "manager_facsimile_url": _facsimile_url(manager_employee),
        "storekeeper_facsimile_url": _facsimile_url(storekeeper_employee),
    }
    return render(request, "orders/mx1_print.html", ctx)


class OrdersHomeView(RoleRequiredMixin, TemplateView):
    template_name = 'orders/index.html'
    allowed_roles = ("manager", "storekeeper", "head_manager", "director", "admin")

    def dispatch(self, request, *args, **kwargs):
        client_agency = _client_agency_from_request(request)
        if client_agency:
            request._client_agency = client_agency
            return TemplateView.dispatch(self, request, *args, **kwargs)
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _next_order_number(order_type: str = "receiving") -> str:
        last = (
            OrderAuditEntry.objects.filter(order_type=order_type)
            .order_by("-created_at")
            .first()
        )
        if not last or not last.order_id:
            return "1"
        match = re.search(r"(\d+)$", last.order_id)
        if not match:
            count = OrderAuditEntry.objects.filter(order_type=order_type).count()
            return str(count + 1)
        return str(int(match.group(1)) + 1)

    def get(self, request, *args, **kwargs):
        status = (request.GET.get("status") or "").lower()
        ok = request.GET.get("ok") == "1"
        submitted = kwargs.get("submitted") or (ok and status != "draft")
        draft_saved = ok and status == "draft"
        error = kwargs.get("error")
        ctx = self.get_context_data(
            submitted=submitted,
            draft_saved=draft_saved,
            error=error,
            **kwargs,
        )
        return self.render_to_response(ctx)

    def post(self, request, *args, **kwargs):
        active_tab = kwargs.get("tab", "journal")
        if active_tab == "packing":
            return self._submit_packing(request)
        if active_tab != "receiving":
            return redirect("/orders/")

        client_agency = getattr(request, "_client_agency", None) or _client_agency_from_request(request)
        if client_agency:
            agency = client_agency
        else:
            agency_id = request.POST.get("agency_id")
            agency = Agency.objects.filter(pk=agency_id).first()
        if not agency:
            return self.get(request, error="Выберите клиента.")

        template_items = []
        template_files = []
        template_files.extend(request.FILES.getlist("template_files"))
        sku_template_file = request.FILES.get("template_sku_file")
        receiving_template_file = request.FILES.get("template_receiving_file")
        if sku_template_file:
            template_files.append(sku_template_file)
        if receiving_template_file:
            template_files.append(receiving_template_file)
        template_names = [f.name for f in template_files if getattr(f, "name", "")]
        if template_files:
            try:
                with transaction.atomic():
                    template_items, template_names = _process_template_uploads(template_files, agency)
            except ValueError as exc:
                return self.get(request, error=str(exc))

        eta_raw = (request.POST.get("eta_at") or "").strip()
        if not eta_raw:
            return self.get(request, error="Укажите плановую дату и время прибытия.")
        try:
            eta_value = datetime.fromisoformat(eta_raw)
        except ValueError:
            return self.get(request, error="Некорректная дата/время прибытия.")
        if timezone.is_naive(eta_value):
            eta_value = timezone.make_aware(eta_value, timezone.get_current_timezone())
        eta_value = timezone.localtime(eta_value)
        min_eta = _min_receiving_eta(timezone.localtime())
        if eta_value < min_eta:
            return self.get(
                request,
                error="Плановая дата/время не может быть раньше следующего дня. После 18:00 — не раньше 13:00.",
            )

        expected_boxes_raw = (request.POST.get("expected_boxes") or "").strip()
        place_type = (request.POST.get("place_type") or "").strip()
        vehicle_number = (request.POST.get("vehicle_number") or "").strip()
        driver_phone = (request.POST.get("driver_phone") or "").strip()
        try:
            expected_boxes_value = int(expected_boxes_raw)
        except (TypeError, ValueError):
            expected_boxes_value = 0
        if expected_boxes_value <= 0:
            return self.get(request, error="Укажите количество мест.")
        if not place_type:
            return self.get(request, error="Выберите тип мест.")
        if not vehicle_number:
            return self.get(request, error="Укажите номер авто.")
        if not _is_valid_driver_phone(driver_phone):
            return self.get(request, error="Введите корректный телефон водителя.")

        sku_codes = request.POST.getlist("sku_code[]")
        sku_ids = request.POST.getlist("sku_id[]")
        names = request.POST.getlist("item_name[]")
        sizes = request.POST.getlist("size[]")
        qtys = request.POST.getlist("qty[]")
        position_comments = request.POST.getlist("position_comment[]")
        items = []
        row_count = max(
            len(sku_codes),
            len(qtys),
            len(names),
            len(position_comments),
            len(sku_ids),
            len(sizes),
        )
        for idx in range(row_count):
            sku_code = sku_codes[idx] if idx < len(sku_codes) else ""
            qty = qtys[idx] if idx < len(qtys) else ""
            if not sku_code and not qty:
                continue
            items.append(
                {
                    "sku_id": sku_ids[idx] if idx < len(sku_ids) else "",
                    "sku_code": sku_code,
                    "name": names[idx] if idx < len(names) else "",
                    "size": sizes[idx] if idx < len(sizes) else "",
                    "qty": qty,
                    "comment": position_comments[idx] if idx < len(position_comments) else "",
                }
            )
        if template_items:
            items = _merge_items(items + template_items)

        submit_action = request.POST.get("submit_action")
        edit_order_id = request.POST.get("edit_order_id")
        role = get_request_role(request)
        previous_entries = []
        can_client_edit = False
        if edit_order_id:
            previous_entries = list(
                OrderAuditEntry.objects.filter(
                    order_id=edit_order_id,
                    order_type="receiving",
                ).order_by("created_at")
            )
            can_client_edit = bool(client_agency and _can_client_edit_draft(previous_entries, client_agency))
            if role != "manager" and not can_client_edit:
                return self.get(request, error="Недостаточно прав для исправления заявки")
        old_payload = {}
        if edit_order_id:
            old_payload = _latest_payload_from_entries(previous_entries)
        status_value = "draft" if submit_action == "draft" else "sent_unconfirmed"
        status_label = "Черновик" if status_value == "draft" else "Ждет подтверждения"
        if edit_order_id and old_payload and submit_action != "send":
            preserved_status = old_payload.get("status") or old_payload.get("submit_action")
            preserved_label = old_payload.get("status_label")
            if preserved_status:
                status_value = preserved_status
            if preserved_label:
                status_label = preserved_label
        payload = {
            "eta_at": eta_raw,
            "expected_boxes": expected_boxes_raw,
            "place_type": place_type,
            "vehicle_number": vehicle_number,
            "driver_phone": driver_phone,
            "comment": request.POST.get("comment"),
            "submit_action": submit_action,
            "status": status_value,
            "status_label": status_label,
            "items": items,
            "documents": [f.name for f in request.FILES.getlist("documents")],
            "template_files": template_names,
        }
        if edit_order_id:
            changes = _describe_payload_changes(old_payload, payload)
            if changes:
                description = f"Исправление заявки №{edit_order_id}: " + "; ".join(changes)
            else:
                description = f"Исправление заявки №{edit_order_id}: без изменений"
            if submit_action == "send":
                description = f"Отправлено менеджеру. {description}"
            log_order_action(
                "update",
                order_id=edit_order_id,
                order_type="receiving",
                user=request.user if request.user.is_authenticated else None,
                agency=agency,
                description=description,
                payload=payload,
            )
            if submit_action == "send" and self.request.GET.get("client"):
                _create_manager_task(edit_order_id, agency, request, timezone.localtime())
            return redirect(f"/orders/receiving/{edit_order_id}/")

        action_label = "черновик" if submit_action == "draft" else "заявка"
        order_id = self._next_order_number(order_type="receiving")
        log_order_action(
            "create",
            order_id=order_id,
            order_type="receiving",
            user=request.user if request.user.is_authenticated else None,
            agency=agency,
            description=f"Заявка на приемку №{order_id} ({action_label})",
            payload=payload,
        )
        if submit_action != "draft" and self.request.GET.get("client"):
            _create_manager_task(order_id, agency, request, timezone.localtime())
        return redirect(
            f"/orders/receiving/?client={agency.id}&ok=1&status={status_value}&order={order_id}"
        )

    def _submit_packing(self, request):
        client_agency = getattr(request, "_client_agency", None) or _client_agency_from_request(request)
        if client_agency:
            agency = client_agency
        else:
            agency_id = request.POST.get("agency_id")
            agency = Agency.objects.filter(pk=agency_id).first()
        if not agency:
            return self.get(request, error="Выберите клиента.")

        email = (request.POST.get("email") or "").strip()
        fio = (request.POST.get("fio") or "").strip()
        plan_date = (request.POST.get("plan_date") or "").strip()
        total_qty_raw = (request.POST.get("total_qty") or "").strip()
        if not email:
            return self.get(request, error="Укажите email.")
        if not fio:
            return self.get(request, error="Укажите ФИО.")
        if not plan_date:
            return self.get(request, error="Укажите плановую дату выполнения.")
        try:
            total_qty_value = int(total_qty_raw)
        except (TypeError, ValueError):
            total_qty_value = 0
        if total_qty_value <= 0:
            return self.get(request, error="Укажите общее количество к обработке.")

        marketplaces = request.POST.getlist("mp[]")
        mp_other = (request.POST.get("mp_other") or "").strip()
        if not marketplaces and not mp_other:
            return self.get(request, error="Выберите маркетплейс или укажите другое.")

        payload = {
            "email": email,
            "fio": fio,
            "org": request.POST.get("org"),
            "plan_date": plan_date,
            "marketplaces": marketplaces,
            "mp_other": request.POST.get("mp_other"),
            "subject": request.POST.get("subject"),
            "total_qty": total_qty_raw,
            "box_mode": request.POST.get("box_mode"),
            "box_mode_other": request.POST.get("box_mode_other"),
            "tasks": request.POST.getlist("tasks[]"),
            "tasks_other": request.POST.get("tasks_other"),
            "marking": request.POST.get("marking"),
            "marking_other": request.POST.get("marking_other"),
            "ship_as": request.POST.get("ship_as"),
            "ship_other": request.POST.get("ship_other"),
            "has_distribution": request.POST.get("has_distribution"),
            "comments": request.POST.get("comments"),
            "files_report": [f.name for f in request.FILES.getlist("files_report")],
            "files_distribution": [f.name for f in request.FILES.getlist("files_distribution")],
            "files_cz": [f.name for f in request.FILES.getlist("files_cz")],
        }
        status_value = "sent_unconfirmed"
        status_label = "Ждет подтверждения"
        payload["status"] = status_value
        payload["status_label"] = status_label
        payload["submit_action"] = "submitted"
        order_id = self._next_order_number(order_type="packing")
        log_order_action(
            "create",
            order_id=order_id,
            order_type="packing",
            user=request.user if request.user.is_authenticated else None,
            agency=agency,
            description=f"Заявка на упаковку №{order_id}",
            payload=payload,
        )
        return redirect(
            f"/orders/packing/?client={agency.id}&ok=1&status={status_value}&order={order_id}"
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        active_tab = kwargs.get("tab", "journal")
        ctx["active_tab"] = active_tab
        ctx["submitted"] = kwargs.get("submitted", False)
        ctx["draft_saved"] = kwargs.get("draft_saved", False)
        ctx["error"] = kwargs.get("error")
        ctx["status_label"] = ctx.get("status_label", "Подготовка заявки")
        ctx["order_number"] = ctx.get("order_number", "")
        ctx["cabinet_url"] = resolve_cabinet_url(get_request_role(self.request))

        client_id = self.request.GET.get("client")
        agency_id = self.request.GET.get("agency")
        agency_key = client_id or agency_id
        client_agency = getattr(self.request, "_client_agency", None) or _client_agency_from_request(self.request)
        agency = client_agency or (Agency.objects.filter(pk=agency_key).first() if agency_key else None)
        client_view = bool(client_agency)
        ctx["agency"] = agency
        ctx["client_view"] = client_view

        if active_tab == "journal":
            raw_entries = OrderAuditEntry.objects.select_related("user", "agency").order_by("-created_at")
            if agency:
                raw_entries = raw_entries.filter(agency=agency)
            raw_entries = list(raw_entries)
            manager_label = _manager_label()
            storekeeper_label = _storekeeper_label()
            receiving_totals = {}
            for entry in raw_entries:
                if entry.order_type != "receiving":
                    continue
                payload = entry.payload or {}
                if payload.get("act") != "receiving":
                    continue
                if entry.order_id in receiving_totals:
                    continue
                total = 0
                for item in payload.get("act_items") or []:
                    qty = _parse_qty_value(item.get("actual_qty"))
                    if qty is None:
                        qty = _parse_qty_value(item.get("qty")) or 0
                    total += qty
                receiving_totals[entry.order_id] = total
            latest_by_order = {}
            status_by_order = {}
            for entry in raw_entries:
                if entry.order_id not in latest_by_order:
                    latest_by_order[entry.order_id] = entry
                if entry.order_id not in status_by_order and _is_status_entry(entry):
                    status_by_order[entry.order_id] = entry
            latest_entries = [
                status_by_order.get(order_id, latest_entry)
                for order_id, latest_entry in latest_by_order.items()
            ]
            latest_entries.sort(key=lambda item: item.created_at, reverse=True)
            if not client_view:
                latest_entries = [entry for entry in latest_entries if not _is_draft_entry(entry)]
            entries_payload = []
            for entry in latest_entries:
                if client_view and entry.order_type == "receiving" and _is_draft_entry(entry) and entry.agency_id:
                    detail_url = f"/orders/receiving/?client={entry.agency_id}&edit={entry.order_id}"
                else:
                    detail_url = f"/orders/{entry.order_type}/{entry.order_id}/"
                    if client_view and entry.agency:
                        detail_url += f"?client={entry.agency.id}"
                entries_payload.append(
                        {
                            "created_at": entry.created_at,
                            "order_type": entry.order_type,
                            "type_label": _order_type_label(entry.order_type),
                            "order_id": entry.order_id,
                            "action_label": _journal_action_label(entry, manager_label, storekeeper_label),
                            "status_label": _journal_status_label(entry),
                            "agency": entry.agency,
                            "actual_qty": receiving_totals.get(entry.order_id),
                            "client_label": _shorten_ip_name(
                                (entry.agency.agn_name or str(entry.agency)) if entry.agency else "-"
                            ),
                            "detail_url": detail_url,
                    }
                )
            ctx["entries"] = entries_payload
        if active_tab == "receiving":
            ctx["agencies"] = Agency.objects.order_by("agn_name")
            ctx["current_time"] = timezone.localtime()
            ctx["min_past_hours"] = 0
            status = self.request.GET.get("status")
            status_label = "Подготовка заявки"
            if status == "draft":
                status_label = "Черновик"
            elif status == "sent_unconfirmed":
                status_label = "Ждет подтверждения"
            elif status in {"warehouse", "on_warehouse"}:
                status_label = "В ожидании поставки товара"
            ctx["status_label"] = status_label
            ctx["order_number"] = self.request.GET.get("order", "")

            edit_order_id = self.request.GET.get("edit")
            if edit_order_id:
                role = get_request_role(self.request)
                entries = list(
                    OrderAuditEntry.objects.filter(
                        order_id=edit_order_id,
                        order_type="receiving",
                    )
                    .select_related("agency")
                    .order_by("created_at")
                )
                can_client_edit = client_view and _can_client_edit_draft(entries, client_agency)
                if _act_storekeeper_signed(entries):
                    ctx["error"] = "Заявка подписана кладовщиком, редактирование запрещено"
                elif role != "manager" and not can_client_edit:
                    ctx["error"] = "Недостаточно прав для исправления заявки"
                elif entries:
                    edit_payload = _latest_payload_from_entries(entries)
                    if not agency:
                        agency = entries[-1].agency
                        ctx["agency"] = agency
                    ctx["edit_mode"] = True
                    ctx["edit_order_id"] = edit_order_id
                    ctx["edit_payload"] = edit_payload
                    ctx["edit_is_draft"] = bool(can_client_edit)
                    ctx["status_label"] = "Черновик" if can_client_edit else "Исправление заявки"
                    ctx["order_number"] = edit_order_id

            sku_options = []
            if agency:
                skus = (
                    SKU.objects.filter(agency=agency, deleted=False)
                    .prefetch_related("barcodes")
                    .order_by("sku_code")
                )
                for sku in skus:
                    barcodes = [barcode.value for barcode in sku.barcodes.all()]
                    size_map = {}
                    for barcode in sku.barcodes.all():
                        size_value = (barcode.size or "").strip()
                        if not size_value:
                            continue
                        size_map.setdefault(size_value, []).append(barcode.value)
                    sku_options.append(
                        {
                            "id": sku.id,
                            "code": sku.sku_code,
                            "name": sku.name,
                            "barcodes_joined": "|".join(barcodes),
                            "sizes_json": json.dumps(size_map, ensure_ascii=True),
                        }
                    )
            ctx["sku_options"] = sku_options
        if active_tab == "packing":
            status = self.request.GET.get("status")
            status_label = "Подготовка заявки"
            if status == "draft":
                status_label = "Черновик"
            elif status in {"sent_unconfirmed", "send", "submitted"}:
                status_label = "Ждет подтверждения"
            ctx["status_label"] = status_label
            ctx["order_number"] = self.request.GET.get("order", "")
        return ctx


class OrdersPackingView(OrdersHomeView):
    template_name = "orders/packing.html"


class OrdersDetailView(RoleRequiredMixin, TemplateView):
    template_name = "orders/detail.html"
    order_type = "receiving"
    allowed_roles = ("manager", "storekeeper", "head_manager", "director", "admin")

    def dispatch(self, request, *args, **kwargs):
        client_agency = _client_agency_from_request(request)
        if client_agency:
            request._client_agency = client_agency
            return TemplateView.dispatch(self, request, *args, **kwargs)
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _payload_from_entries(entries):
        for entry in reversed(entries):
            payload = entry.payload or {}
            if not payload:
                continue
            significant_keys = set(payload.keys()) - {
                "comment",
                "message",
                "status",
                "status_label",
                "submit_action",
            }
            if significant_keys:
                return payload
        return entries[-1].payload or {} if entries else {}

    def post(self, request, *args, **kwargs):
        order_id = kwargs.get("order_id")
        if not order_id:
            return redirect("/orders/")
        action = request.POST.get("action")
        if action == "send_to_warehouse":
            latest = (
                OrderAuditEntry.objects.filter(order_id=order_id, order_type=self.order_type)
                .select_related("agency")
                .order_by("-created_at")
                .first()
            )
            if not latest:
                return redirect("/orders/")
            payload = dict(latest.payload or {})
            payload["status"] = "warehouse"
            payload["status_label"] = "В ожидании поставки товара"
            log_order_action(
                "status",
                order_id=order_id,
                order_type=self.order_type,
                user=request.user if request.user.is_authenticated else None,
                agency=latest.agency if latest else None,
                description="Подтверждено и отправлено на склад",
                payload=payload,
            )
            Task.objects.filter(
                route=f"/orders/receiving/{order_id}/",
                assigned_to__role="manager",
            ).exclude(status="done").update(status="done")
            observer = Employee.objects.filter(
                user=request.user, is_active=True
            ).first()
            _create_storekeeper_task(
                order_id,
                latest.agency,
                request,
                timezone.localtime(),
                observer=observer,
            )
            client_param = request.GET.get("client")
            suffix = f"?client={client_param}" if client_param else ""
            return redirect(f"/orders/{self.order_type}/{order_id}/{suffix}")
        if action == "send_act_to_client":
            entries = list(
                OrderAuditEntry.objects.filter(order_id=order_id, order_type=self.order_type)
                .select_related("agency")
                .order_by("created_at")
            )
            if not entries:
                return redirect("/orders/")
            role = get_request_role(request)
            if role not in {"manager", "head_manager", "director", "admin"}:
                return redirect("/orders/")
            _send_act_to_client(order_id, entries, request.user)
            client_param = request.GET.get("client")
            suffix = f"?client={client_param}" if client_param else ""
            return redirect(f"/orders/{self.order_type}/{order_id}/{suffix}")
        if action == "create_receiving_act":
            return redirect(f"/orders/{self.order_type}/{order_id}/act/")
        comment = (request.POST.get("comment") or "").strip()
        if comment:
            latest = (
                OrderAuditEntry.objects.filter(order_id=order_id, order_type=self.order_type)
                .select_related("agency")
                .order_by("-created_at")
                .first()
            )
            log_order_action(
                "comment",
                order_id=order_id,
                order_type=self.order_type,
                user=request.user if request.user.is_authenticated else None,
                agency=latest.agency if latest else None,
                description=comment,
                payload={"comment": comment},
            )
        client_param = request.GET.get("client")
        suffix = f"?client={client_param}" if client_param else ""
        return redirect(f"/orders/{self.order_type}/{order_id}/{suffix}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        order_id = kwargs.get("order_id")
        client_agency = getattr(self.request, "_client_agency", None) or _client_agency_from_request(self.request)
        client_view = bool(client_agency)
        entries = (
            OrderAuditEntry.objects.filter(order_id=order_id, order_type=self.order_type)
            .select_related("user", "agency")
            .order_by("created_at")
        )
        entries_list = list(entries)
        latest = entries_list[-1] if entries_list else None
        status_entry = _current_status_entry(entries_list)
        payload = self._payload_from_entries(entries_list)
        ctx["order_id"] = order_id
        ctx["order_type"] = self.order_type
        ctx["order_title"] = _order_title_label(self.order_type, payload)
        ctx["agency"] = latest.agency if latest else client_agency
        ctx["client_view"] = client_view
        status_text = _status_label_from_entry(status_entry) if status_entry else "-"
        ctx["status_label"] = status_text
        ctx["responsible"] = _current_responsible_label(status_entry)
        ctx["cabinet_url"] = resolve_cabinet_url(get_request_role(self.request))
        can_send_to_warehouse = False
        role = get_request_role(self.request)
        signed_by_storekeeper = _act_storekeeper_signed(entries_list)
        signed_by_manager = _act_manager_signed(entries_list)
        can_edit_order = bool(role == "manager" and not client_view)
        if "товар принят" in (status_text or "").lower():
            can_edit_order = False
        if signed_by_storekeeper or signed_by_manager:
            can_edit_order = False
        ctx["can_edit_order"] = can_edit_order
        can_create_receiving_act = False
        if status_entry and not client_view and role == "manager":
            payload = status_entry.payload or {}
            status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
            status_label = (payload.get("status_label") or "").lower()
            can_send_to_warehouse = status_value in {"sent_unconfirmed", "send", "submitted"}
            if "склад" in status_label or "ожидании поставки" in status_label or status_value in {"warehouse", "on_warehouse"}:
                can_send_to_warehouse = False
            if status_value == "draft":
                can_send_to_warehouse = False
        if status_entry and not client_view and role == "storekeeper":
            payload = status_entry.payload or {}
            status_value = (payload.get("status") or payload.get("submit_action") or "").lower()
            status_label = (payload.get("status_label") or "").lower()
            has_receiving_act = any(
                (entry.payload or {}).get("act") == "receiving" for entry in entries_list
            )
            if not has_receiving_act and (
                status_value in {"warehouse", "on_warehouse"}
                or "склад" in status_label
                or "ожидании поставки" in status_label
            ):
                can_create_receiving_act = True
        ctx["can_send_to_warehouse"] = can_send_to_warehouse
        ctx["can_create_receiving_act"] = can_create_receiving_act
        client_label = "-"
        if latest and latest.agency:
            name = latest.agency.agn_name or latest.agency.fio_agn or str(latest.agency)
            client_label = _shorten_ip_name(name)
        ctx["client_label"] = client_label
        ctx["meta"] = {
            "eta_at": _format_datetime_value(payload.get("eta_at")),
            "expected_boxes": payload.get("expected_boxes"),
            "place_type": _place_type_label(payload.get("place_type")),
            "vehicle_number": payload.get("vehicle_number"),
            "driver_phone": payload.get("driver_phone"),
            "comment": payload.get("comment"),
        }
        act_entry_for_items = _find_act_entry(entries_list, "receiving", "акт приемки")
        act_items = (act_entry_for_items.payload or {}).get("act_items") if act_entry_for_items else []
        items = payload.get("items") or []
        display_items = []
        if items:
            actual_map = {}
            act_item_by_key = {}
            for act_item in act_items:
                key = _item_key(act_item.get("sku_code"), act_item.get("name"), act_item.get("size"))
                qty = _parse_qty_value(act_item.get("actual_qty"))
                if qty is None:
                    qty = _parse_qty_value(act_item.get("qty")) or 0
                actual_map[key] = (actual_map.get(key) or 0) + qty
                if key not in act_item_by_key:
                    act_item_by_key[key] = act_item
            planned_keys = set()
            for item in items:
                key = _item_key(item.get("sku_code"), item.get("name"), item.get("size"))
                planned_keys.add(key)
                display_items.append(
                    {
                        "sku_code": item.get("sku_code"),
                        "name": item.get("name"),
                        "size": item.get("size"),
                        "qty": item.get("qty"),
                        "comment": item.get("comment"),
                        "actual_qty": actual_map.get(key),
                    }
                )
            for key, act_item in act_item_by_key.items():
                if key in planned_keys:
                    continue
                planned_qty = act_item.get("planned_qty")
                if planned_qty in (None, ""):
                    planned_qty = act_item.get("qty")
                actual_qty = actual_map.get(key)
                if actual_qty is None:
                    actual_qty = _parse_qty_value(act_item.get("actual_qty"))
                    if actual_qty is None:
                        actual_qty = _parse_qty_value(act_item.get("qty"))
                display_items.append(
                    {
                        "sku_code": act_item.get("sku_code"),
                        "name": act_item.get("name"),
                        "size": act_item.get("size"),
                        "qty": planned_qty if planned_qty not in (None, "") else None,
                        "comment": act_item.get("comment"),
                        "actual_qty": actual_qty,
                    }
                )
        elif act_items:
            for item in act_items:
                planned_qty = item.get("planned_qty")
                if planned_qty in (None, ""):
                    planned_qty = item.get("qty")
                actual_qty = _parse_qty_value(item.get("actual_qty"))
                display_items.append(
                    {
                        "sku_code": item.get("sku_code"),
                        "name": item.get("name"),
                        "size": item.get("size"),
                        "qty": planned_qty,
                        "comment": item.get("comment"),
                        "actual_qty": actual_qty,
                    }
                )
        ctx["items"] = display_items
        participants = []
        seen = set()
        for entry in entries_list:
            label = _actor_label(entry.user, entry.agency, client_view=client_view)
            if label in seen:
                continue
            seen.add(label)
            participants.append(label)
        ctx["participants"] = participants
        comments = [
            {
                "created_at": entry.created_at,
                "actor_label": _actor_label(entry.user, entry.agency, client_view=client_view),
                "text": entry.description or (entry.payload or {}).get("comment") or "-",
            }
            for entry in entries_list
            if entry.action == "comment"
        ]
        comments.reverse()
        ctx["comments"] = comments
        ctx["history"] = [
            {
                "created_at": entry.created_at,
                "action_label": _history_action_label(entry),
                "status_label": _status_label_from_entry(entry),
                "description": _format_message_text(entry.description),
                "actor_label": _history_actor_label(
                    entry,
                    client_view=client_view,
                    client_label=client_label,
                ),
            }
            for entry in reversed(entries_list)
            if entry.action != "comment"
        ]
        act_entry = _find_act_entry(entries_list, "receiving", "акт приемки")
        placement_entry = _find_act_entry(entries_list, "placement", "акт размещения")
        ctx["act_entry"] = act_entry
        ctx["placement_act_entry"] = placement_entry
        ctx["has_receiving_act"] = bool(act_entry)
        ctx["has_placement_act"] = bool(placement_entry)
        if act_entry:
            ctx["act_label"] = (act_entry.payload or {}).get("act_label") or "Акт приемки"
        else:
            ctx["act_label"] = ""
        if placement_entry:
            ctx["placement_act_label"] = (placement_entry.payload or {}).get("act_label") or "Акт размещения"
        else:
            ctx["placement_act_label"] = ""
        role = get_request_role(self.request)
        ctx["can_create_placement_act"] = bool(
            role == "storekeeper" and act_entry and not placement_entry
        )
        ctx["can_send_act_to_client"] = bool(
            role in {"manager", "head_manager", "director", "admin"}
            and act_entry
            and placement_entry
            and not _is_done_status(status_entry)
            and _act_manager_signed_from_payload(act_entry.payload or {})
            and not client_view
        )
        return ctx


class PackingDetailView(OrdersDetailView):
    order_type = "packing"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        order_id = kwargs.get("order_id")
        entries_list = list(
            OrderAuditEntry.objects.filter(order_id=order_id, order_type=self.order_type)
            .select_related("user", "agency")
            .order_by("created_at")
        )
        payload = self._payload_from_entries(entries_list)
        marketplaces = payload.get("marketplaces") or []
        if isinstance(marketplaces, str):
            marketplaces = [marketplaces] if marketplaces else []
        mp_other = (payload.get("mp_other") or "").strip()
        if mp_other:
            marketplaces = list(marketplaces) + [mp_other]
        tasks = payload.get("tasks") or []
        if isinstance(tasks, str):
            tasks = [tasks] if tasks else []
        tasks_other = (payload.get("tasks_other") or "").strip()
        if tasks_other:
            tasks = list(tasks) + [tasks_other]
        box_mode = _resolve_choice_value(payload.get("box_mode"), payload.get("box_mode_other"))
        marking = _resolve_choice_value(payload.get("marking"), payload.get("marking_other"))
        ship_as = _resolve_choice_value(payload.get("ship_as"), payload.get("ship_other"))
        ctx["packing_fields"] = [
            {"label": "Email", "value": _format_payload_value(payload.get("email"))},
            {"label": "ФИО", "value": _format_payload_value(payload.get("fio"))},
            {"label": "Организация", "value": _format_payload_value(payload.get("org"))},
            {"label": "Плановая дата", "value": _format_payload_value(payload.get("plan_date"))},
            {"label": "Маркетплейсы", "value": _format_payload_list(marketplaces)},
            {"label": "Предмет", "value": _format_payload_value(payload.get("subject"))},
            {"label": "Общее количество", "value": _format_payload_value(payload.get("total_qty"))},
            {"label": "Кратность коробов", "value": _format_payload_value(box_mode)},
            {"label": "Работы", "value": _format_payload_list(tasks)},
            {"label": "Маркировка", "value": _format_payload_value(marking)},
            {"label": "Отгрузка", "value": _format_payload_value(ship_as)},
            {"label": "Распределение", "value": _format_payload_value(payload.get("has_distribution"))},
            {"label": "Комментарий", "value": _format_payload_value(payload.get("comments"))},
            {"label": "Файлы отчета", "value": _format_payload_list(payload.get("files_report"))},
            {"label": "Файл распределения", "value": _format_payload_list(payload.get("files_distribution"))},
            {"label": "Файлы ЧЗ", "value": _format_payload_list(payload.get("files_cz"))},
        ]
        ctx["packing_meta"] = {
            "plan_date": _format_payload_value(payload.get("plan_date")),
            "total_qty": _format_payload_value(payload.get("total_qty")),
            "marketplaces": _format_payload_list(marketplaces),
            "box_mode": _format_payload_value(box_mode),
            "marking": _format_payload_value(marking),
            "ship_as": _format_payload_value(ship_as),
            "has_distribution": _format_payload_value(payload.get("has_distribution")),
            "comment": _format_payload_value(payload.get("comments")),
        }
        ctx["items"] = []
        ctx["can_edit_order"] = False
        ctx["can_send_to_warehouse"] = False
        ctx["can_create_receiving_act"] = False
        ctx["has_receiving_act"] = False
        ctx["has_placement_act"] = False
        ctx["act_label"] = ""
        ctx["placement_act_label"] = ""
        ctx["can_send_act_to_client"] = False
        return ctx


class ReceivingActView(RoleRequiredMixin, TemplateView):
    template_name = "orders/receiving_act.html"
    order_type = "receiving"
    allowed_roles = ("storekeeper", "manager", "head_manager", "director", "admin")

    def dispatch(self, request, *args, **kwargs):
        client_agency = _client_agency_from_request(request)
        if client_agency:
            order_id = kwargs.get("order_id")
            entries = self._load_entries(order_id)
            if not entries:
                return HttpResponseForbidden("Доступ запрещен")
            if not any(entry.agency_id == client_agency.id for entry in entries if entry.agency_id):
                return HttpResponseForbidden("Доступ запрещен")
            if not self._act_entry(entries):
                return HttpResponseForbidden("Доступ запрещен")
            if request.method != "GET":
                return HttpResponseForbidden("Доступ запрещен")
            status_entry = _current_status_entry(entries)
            if status_entry:
                payload = dict(status_entry.payload or {})
                if payload.get("act_sent") and not payload.get("act_viewed"):
                    payload["status"] = payload.get("status") or "done"
                    payload["status_label"] = payload.get("status_label") or "Выполнена"
                    payload["act_viewed"] = True
                    payload["act_viewed_at"] = timezone.localtime().isoformat()
                    log_order_action(
                        "status",
                        order_id=order_id,
                        order_type=self.order_type,
                        user=request.user if request.user.is_authenticated else None,
                        agency=entries[-1].agency if entries else None,
                        description="Акт приемки просмотрен клиентом",
                        payload=payload,
                    )
            response = request.GET.get("response")
            raw_return_url = request.GET.get("return") or request.META.get("HTTP_REFERER")
            return_url = _safe_return_url(request, raw_return_url, request.path)
            return redirect(_client_print_url(order_id, client_agency, response, return_url))
        return super().dispatch(request, *args, **kwargs)

    def _load_entries(self, order_id):
        return list(
            OrderAuditEntry.objects.filter(order_id=order_id, order_type=self.order_type)
            .select_related("user", "agency")
            .order_by("created_at")
        )

    def _act_entry(self, entries):
        return _act_entry_from_entries(entries, "receiving")

    def _can_create(self, entries):
        if not entries:
            return False
        if self._act_entry(entries):
            return False
        status_entry = _current_status_entry(entries)
        return _warehouse_status_from_entry(status_entry)

    def get(self, request, *args, **kwargs):
        ok = request.GET.get("ok") == "1"
        error = request.GET.get("error")
        ctx = self.get_context_data(ok=ok, error=error, **kwargs)
        return self.render_to_response(ctx)

    def post(self, request, *args, **kwargs):
        order_id = kwargs.get("order_id")
        entries = self._load_entries(order_id)
        if not entries:
            return redirect("/orders/")
        role = get_request_role(request)
        if role != "storekeeper":
            return HttpResponseForbidden("Доступ запрещен")
        if not self._can_create(entries):
            return redirect(f"/orders/receiving/{order_id}/act/?error=1")
        payload = _latest_payload_from_entries(entries)
        items = payload.get("items") or []
        has_planned_items = bool(items)
        actual_raw = request.POST.getlist("actual_qty[]")
        eta_raw = (request.POST.get("eta_at") or "").strip()
        vehicle_number = (request.POST.get("vehicle_number") or "").strip()
        if not eta_raw:
            return redirect(f"/orders/receiving/{order_id}/act/?error=1")
        try:
            eta_value = datetime.fromisoformat(eta_raw)
        except ValueError:
            return redirect(f"/orders/receiving/{order_id}/act/?error=1")
        if timezone.is_naive(eta_value):
            eta_value = timezone.make_aware(eta_value, timezone.get_current_timezone())
        eta_value = timezone.localtime(eta_value)
        if not vehicle_number:
            return redirect(f"/orders/receiving/{order_id}/act/?error=1")
        act_items = []
        has_mismatch = False
        for idx, item in enumerate(items):
            raw_value = actual_raw[idx] if idx < len(actual_raw) else ""
            actual_value = _parse_qty_value(raw_value)
            if actual_value is None:
                return redirect(f"/orders/receiving/{order_id}/act/?error=1")
            planned_value = _parse_qty_value(item.get("qty")) or 0
            if planned_value != actual_value:
                has_mismatch = True
            act_items.append(
                {
                    "sku_code": item.get("sku_code"),
                    "name": item.get("name"),
                    "size": item.get("size"),
                    "planned_qty": item.get("qty"),
                    "actual_qty": actual_value,
                    "comment": item.get("comment"),
                }
            )
        extra_sku_codes = request.POST.getlist("extra_sku_code[]")
        extra_barcodes = request.POST.getlist("extra_barcode[]")
        extra_names = request.POST.getlist("extra_name[]")
        extra_sizes = request.POST.getlist("extra_size[]")
        extra_planned = request.POST.getlist("extra_planned_qty[]")
        extra_actual = request.POST.getlist("extra_actual_qty[]")
        extra_count = max(
            len(extra_names),
            len(extra_sizes),
            len(extra_planned),
            len(extra_actual),
            len(extra_sku_codes),
            len(extra_barcodes),
        )
        for idx in range(extra_count):
            sku_code = (extra_sku_codes[idx] if idx < len(extra_sku_codes) else "").strip()
            barcode = (extra_barcodes[idx] if idx < len(extra_barcodes) else "").strip()
            name = (extra_names[idx] if idx < len(extra_names) else "").strip()
            size = (extra_sizes[idx] if idx < len(extra_sizes) else "").strip()
            planned_raw = extra_planned[idx] if idx < len(extra_planned) else ""
            actual_raw_value = extra_actual[idx] if idx < len(extra_actual) else ""
            if not name and not size and not planned_raw and not actual_raw_value:
                continue
            if not name:
                return redirect(f"/orders/receiving/{order_id}/act/?error=1")
            planned_value = _parse_qty_value(planned_raw)
            if planned_value is None:
                planned_value = 0
            actual_value = _parse_qty_value(actual_raw_value)
            if actual_value is None:
                return redirect(f"/orders/receiving/{order_id}/act/?error=1")
            if has_planned_items and planned_value != actual_value:
                has_mismatch = True
            act_items.append(
                {
                    "sku_code": sku_code,
                    "barcode": barcode,
                    "name": name,
                    "size": size,
                    "planned_qty": planned_value,
                    "actual_qty": actual_value,
                    "comment": "",
                    "extra": True,
                }
            )
        if not act_items:
            return redirect(f"/orders/receiving/{order_id}/act/?error=1")
        status_entry = _current_status_entry(entries)
        latest = entries[-1]
        act_payload = dict((status_entry.payload or {}) if status_entry else {})
        if not act_payload.get("status"):
            act_payload["status"] = "warehouse"
        act_payload["status_label"] = (
            "Товар принят на склад с расхождениями" if has_mismatch else "Товар принят на склад"
        )
        act_payload["eta_at"] = eta_value.isoformat()
        act_payload["vehicle_number"] = vehicle_number
        act_payload["act"] = "receiving"
        act_payload["act_label"] = (
            "Акт приемки с расхождениями" if has_mismatch else "Акт приемки"
        )
        act_payload["act_mismatch"] = has_mismatch
        act_payload["act_items"] = act_items
        log_order_action(
            "status",
            order_id=order_id,
            order_type=self.order_type,
            user=request.user if request.user.is_authenticated else None,
            agency=latest.agency,
            description="Создан акт приемки",
            payload=act_payload,
        )
        return redirect(f"/orders/receiving/{order_id}/placement/?ok=1")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        order_id = kwargs.get("order_id")
        entries = self._load_entries(order_id)
        latest = entries[-1] if entries else None
        client_agency = getattr(self.request, "_client_agency", None)
        client_view = bool(client_agency)
        status_entry = _current_status_entry(entries)
        payload = _latest_payload_from_entries(entries)
        order_title = _order_title_label(self.order_type, payload)
        items = payload.get("items") or []
        act_entry = self._act_entry(entries)
        act_items = (act_entry.payload or {}).get("act_items") if act_entry else []
        placement_entry = _find_act_entry(entries, "placement", "акт размещения")
        placement_closed = _placement_closed(entries)
        role = get_request_role(self.request)
        if act_entry:
            act_label = (act_entry.payload or {}).get("act_label") or "Акт приемки"
        else:
            act_label = "Акт приемки"
        act_documents = []
        if act_entry and placement_entry and placement_closed:
            act_documents = [
                {
                    "label": "Акт приемки печатная форма",
                    "url": f"/orders/receiving/{order_id}/act/print/",
                },
                {"label": "МХ-1", "url": f"/orders/receiving/{order_id}/act/mx1/print/"},
            ]
            if role == "storekeeper":
                act_documents = [act_documents[0]]
        can_submit = role == "storekeeper" and self._can_create(entries)
        if client_view:
            can_submit = False
        display_items = []
        base_items = act_items or items
        sku_ids = set()
        sku_codes = set()
        for item in base_items:
            sku_code = str(item.get("sku_code") or "").strip()
            if sku_code:
                sku_codes.add(sku_code)
            sku_id_raw = item.get("sku_id")
            if sku_id_raw:
                try:
                    sku_ids.add(int(sku_id_raw))
                except (TypeError, ValueError):
                    pass
        sku_by_id = {}
        if sku_ids:
            for sku in SKU.objects.filter(id__in=sku_ids, deleted=False).prefetch_related("barcodes"):
                sku_by_id[sku.id] = sku
        sku_by_code = {}
        if sku_codes:
            sku_qs = SKU.objects.filter(sku_code__in=sku_codes, deleted=False)
            if latest and latest.agency_id:
                sku_qs = sku_qs.filter(agency_id=latest.agency_id)
            for sku in sku_qs.prefetch_related("barcodes"):
                sku_by_code.setdefault(sku.sku_code, sku)
        for idx, item in enumerate(base_items):
            actual_value = "" if not act_items else item.get("actual_qty")
            if act_items:
                planned_value = item.get("planned_qty")
                name = item.get("name")
                size = item.get("size")
            else:
                planned_value = item.get("qty")
                name = item.get("name")
                size = item.get("size")
            sku_code_value = item.get("sku_code") or ""
            sku_id_value = item.get("sku_id")
            sku_id = None
            if sku_id_value not in (None, ""):
                try:
                    sku_id = int(sku_id_value)
                except (TypeError, ValueError):
                    sku_id = None
            sku = sku_by_id.get(sku_id) or sku_by_code.get(sku_code_value)
            barcode_value = item.get("barcode") or _barcode_value_for_sku(sku, size)
            display_items.append(
                {
                    "sku_code": sku_code_value or "-",
                    "barcode": barcode_value,
                    "name": name or "-",
                    "size": size or "-",
                    "planned_qty": planned_value if planned_value not in (None, "") else "-",
                    "actual_qty": actual_value if actual_value is not None else "",
                    "comment": item.get("comment") or "",
                }
            )
        client_label = "-"
        if latest and latest.agency:
            name = latest.agency.agn_name or latest.agency.fio_agn or str(latest.agency)
            client_label = _shorten_ip_name(name)
        sku_options = []
        sku_name_options = []
        barcode_options = []
        barcode_map = {}
        if latest and latest.agency_id:
            name_seen = set()
            barcode_seen = set()
            for sku in (
                SKU.objects.filter(agency_id=latest.agency_id, deleted=False)
                .prefetch_related("barcodes")
                .order_by("sku_code")
            ):
                barcode_values = []
                for barcode in sku.barcodes.all():
                    value = (barcode.value or "").strip()
                    if not value:
                        continue
                    barcode_values.append(value)
                    if value not in barcode_seen:
                        barcode_seen.add(value)
                        barcode_options.append(value)
                    if value not in barcode_map:
                        barcode_map[value] = {
                            "sku": sku.sku_code,
                            "name": sku.name,
                            "size": (barcode.size or sku.size or "").strip(),
                        }
                sku_code_barcode = (sku.code or "").strip()
                if sku_code_barcode:
                    if sku_code_barcode not in barcode_values:
                        barcode_values.append(sku_code_barcode)
                    if sku_code_barcode not in barcode_seen:
                        barcode_seen.add(sku_code_barcode)
                        barcode_options.append(sku_code_barcode)
                    barcode_map.setdefault(
                        sku_code_barcode,
                        {
                            "sku": sku.sku_code,
                            "name": sku.name,
                            "size": (sku.size or "").strip(),
                        },
                    )
                sku_options.append(
                    {
                        "code": sku.sku_code,
                        "name": sku.name,
                        "barcodes_joined": "|".join(barcode_values),
                    }
                )
                if sku.name and sku.name not in name_seen:
                    name_seen.add(sku.name)
                    sku_name_options.append(sku.name)
        arrival_value = payload.get("eta_at")
        if act_entry:
            act_payload = act_entry.payload or {}
            arrival_value = act_payload.get("eta_at") or arrival_value
        vehicle_value = payload.get("vehicle_number")
        driver_phone_value = payload.get("driver_phone")
        if act_entry:
            act_payload = act_entry.payload or {}
            vehicle_value = act_payload.get("vehicle_number") or vehicle_value
            driver_phone_value = act_payload.get("driver_phone") or driver_phone_value
        arrival_input = ""
        if arrival_value:
            try:
                parsed = datetime.fromisoformat(str(arrival_value))
                if timezone.is_naive(parsed):
                    parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
                parsed = timezone.localtime(parsed)
                arrival_input = parsed.strftime("%Y-%m-%dT%H:%M")
            except (TypeError, ValueError):
                arrival_input = ""
        cabinet_url = resolve_cabinet_url(get_request_role(self.request))
        if client_view and client_agency:
            cabinet_url = f"/client/dashboard/?client={client_agency.id}"
        driver_phone = _format_payload_value(driver_phone_value)
        ctx.update(
            {
                "order_id": order_id,
                "order_title": order_title,
                "client_label": client_label,
                "status_label": _status_label_from_entry(status_entry) if status_entry else "-",
                "cabinet_url": cabinet_url,
                "client_view": client_view,
                "client_param": client_agency.id if client_agency else "",
                "arrival_at": _format_datetime_value(arrival_value),
                "arrival_value": _format_datetime_value(arrival_value),
                "arrival_input": arrival_input,
                "vehicle_number": vehicle_value or "-",
                "vehicle_value": vehicle_value or "",
                "driver_phone": driver_phone,
                "can_submit": can_submit,
                "can_add_items": can_submit,
                "act_exists": bool(act_entry),
                "act_label": act_label,
                "act_documents": act_documents,
                "items": display_items,
                "sku_options": sku_options,
                "sku_name_options": sku_name_options,
                "barcode_options": barcode_options,
                "barcode_map": barcode_map,
                "agency_id": latest.agency_id if latest else "",
                "ok": kwargs.get("ok", False),
                "error": kwargs.get("error"),
            }
        )
        return ctx


class PlacementActView(RoleRequiredMixin, TemplateView):
    template_name = "orders/placement_act.html"
    order_type = "receiving"
    allowed_roles = ("storekeeper", "head_manager", "director", "admin", "manager")

    def _load_entries(self, order_id):
        return list(
            OrderAuditEntry.objects.filter(order_id=order_id, order_type=self.order_type)
            .select_related("user", "agency")
            .order_by("created_at")
        )

    def _receiving_act_entry(self, entries):
        return _act_entry_from_entries(entries, "receiving")

    def _placement_act_entry(self, entries):
        return _act_entry_from_entries(entries, "placement")

    def _can_create(self, entries):
        if not entries:
            return False
        if not self._receiving_act_entry(entries):
            return False
        return True

    def get(self, request, *args, **kwargs):
        ok = request.GET.get("ok") == "1"
        error = request.GET.get("error")
        ctx = self.get_context_data(ok=ok, error=error, **kwargs)
        return self.render_to_response(ctx)

    def post(self, request, *args, **kwargs):
        order_id = kwargs.get("order_id")
        entries = self._load_entries(order_id)
        if not entries:
            return redirect("/orders/")
        action = (request.POST.get("action") or "close").lower()
        role = get_request_role(request)
        if role != "storekeeper":
            return redirect(f"/orders/receiving/{order_id}/placement/")
        if action == "open" and _act_storekeeper_signed(entries):
            return redirect(f"/orders/receiving/{order_id}/placement/?error=signed")
        if not self._can_create(entries):
            return redirect(f"/orders/receiving/{order_id}/placement/?error=1")
        placement_act = self._placement_act_entry(entries)
        if action == "open":
            if not placement_act:
                return redirect(f"/orders/receiving/{order_id}/placement/")
            payload = placement_act.payload or {}
            current_state = (payload.get("act_state") or "closed").lower()
            if current_state == "open":
                return redirect(f"/orders/receiving/{order_id}/placement/")
            act_payload = dict(payload)
            if not act_payload.get("status"):
                act_payload["status"] = "warehouse"
            act_payload["status_label"] = "Размещение на складе"
            act_payload["act"] = "placement"
            act_payload["act_label"] = act_payload.get("act_label") or "Акт размещения"
            act_payload["act_state"] = "open"
            latest = entries[-1]
            log_order_action(
                "status",
                order_id=order_id,
                order_type=self.order_type,
                user=request.user if request.user.is_authenticated else None,
                agency=latest.agency,
                description="Открыт акт размещения",
                payload=act_payload,
            )
            return redirect(f"/orders/receiving/{order_id}/placement/")
        receiving_act = self._receiving_act_entry(entries)
        act_items = (receiving_act.payload or {}).get("act_items") if receiving_act else []
        if not act_items:
            return redirect(f"/orders/receiving/{order_id}/placement/?error=1")
        placement_items = []
        boxes_raw = request.POST.get("boxes_json") or "[]"
        pallets_raw = request.POST.get("pallets_json") or "[]"
        try:
            boxes_data = json.loads(boxes_raw)
            pallets_data = json.loads(pallets_raw)
        except json.JSONDecodeError:
            return redirect(f"/orders/receiving/{order_id}/placement/?error=1")
        if not isinstance(boxes_data, list):
            boxes_data = []
        if isinstance(pallets_data, list):
            pallets_data = [
                pallet for pallet in pallets_data
                if isinstance(pallet, dict)
                and ((pallet.get("items") or []) or (pallet.get("boxes") or []))
            ]
        else:
            pallets_data = []
        totals = {}
        def add_total(item, qty, field):
            key = _item_key(item.get("sku"), item.get("name"), item.get("size"))
            entry = totals.setdefault(key, {"box": 0, "pallet": 0, "total": 0})
            entry[field] += qty
            entry["total"] += qty
        if isinstance(boxes_data, list):
            for box in boxes_data:
                for item in (box or {}).get("items", []) or []:
                    qty = _parse_qty_value(item.get("qty")) or 0
                    add_total(item, qty, "box")
        if isinstance(pallets_data, list):
            for pallet in pallets_data:
                for item in (pallet or {}).get("items", []) or []:
                    qty = _parse_qty_value(item.get("qty")) or 0
                    add_total(item, qty, "pallet")
        placement_items = []
        for item in act_items:
            key = _item_key(item.get("sku_code"), item.get("name"), item.get("size"))
            entry = totals.get(key, {"box": 0, "pallet": 0, "total": 0})
            actual_qty = _parse_qty_value(item.get("actual_qty")) or 0
            if entry["total"] > actual_qty:
                return redirect(f"/orders/receiving/{order_id}/placement/?error=1")
            if entry["total"] != actual_qty:
                return redirect(f"/orders/receiving/{order_id}/placement/?error=1")
            placement_items.append(
                {
                    "sku_code": item.get("sku_code"),
                    "name": item.get("name"),
                    "size": item.get("size"),
                    "actual_qty": actual_qty,
                    "box_qty": entry["box"],
                    "pallet_qty": entry["pallet"],
                    "comment": item.get("comment"),
                }
            )
        for box in boxes_data:
            if not isinstance(box, dict):
                continue
            if not box.get("sealed"):
                return redirect(f"/orders/receiving/{order_id}/placement/?error=1")
        pallet_box_codes = set()
        for pallet in pallets_data:
            if not isinstance(pallet, dict):
                continue
            for box_code in (pallet.get("boxes") or []):
                if box_code:
                    pallet_box_codes.add(box_code)
        unassigned_boxes = [
            box
            for box in boxes_data
            if isinstance(box, dict)
            and (box.get("code") or "")
            and (box.get("code") not in pallet_box_codes)
        ]
        if unassigned_boxes:
            return redirect(f"/orders/receiving/{order_id}/placement/?error=1")
        default_location = "Поле приемки"
        for pallet in pallets_data:
            if not isinstance(pallet, dict):
                continue
            if not pallet.get("sealed"):
                return redirect(f"/orders/receiving/{order_id}/placement/?error=1")
            location_value = pallet.get("location")
            location_text = ""
            if isinstance(location_value, dict):
                zone = (location_value.get("zone") or "").strip()
                rack = (location_value.get("rack") or pallet.get("rack") or "").strip()
                row = (location_value.get("row") or pallet.get("row") or "").strip()
                section = (location_value.get("section") or "").strip()
                tier = (location_value.get("tier") or "").strip()
                shelf = (location_value.get("shelf") or pallet.get("shelf") or "").strip()
                cell = (location_value.get("cell") or "").strip()
                parts = []
                if zone:
                    parts.append(f"Зона {zone}")
                if rack:
                    parts.append(f"Стеллаж {rack}")
                if row:
                    parts.append(f"Ряд {row}")
                if section:
                    parts.append(f"Секция {section}")
                if tier:
                    parts.append(f"Ярус {tier}")
                if shelf:
                    parts.append(f"Полка {shelf}")
                if cell:
                    parts.append(f"Ячейка {cell}")
                location_text = " · ".join(parts)
            elif isinstance(location_value, str):
                location_text = location_value.strip()
            if not location_text:
                location_text = default_location
            pallet["location"] = location_text
        status_entry = _current_status_entry(entries)
        latest = entries[-1]
        act_payload = dict((status_entry.payload or {}) if status_entry else {})
        if not act_payload.get("status"):
            act_payload["status"] = "warehouse"
        act_payload["status_label"] = "Товар принят и размещен на складе"
        act_payload["act"] = "placement"
        act_payload["act_label"] = "Акт размещения"
        act_payload["act_state"] = "closed"
        act_payload["act_items"] = placement_items
        act_payload["act_boxes"] = boxes_data if isinstance(boxes_data, list) else []
        act_payload["act_pallets"] = pallets_data if isinstance(pallets_data, list) else []
        has_closed_act = any(
            (entry.payload or {}).get("act") == "placement"
            and ((entry.payload or {}).get("act_state") or "closed") == "closed"
            for entry in entries
        )
        log_order_action(
            "status",
            order_id=order_id,
            order_type=self.order_type,
            user=request.user if request.user.is_authenticated else None,
            agency=latest.agency,
            description="Обновлен акт размещения" if has_closed_act else "Создан акт размещения",
            payload=act_payload,
        )
        if not has_closed_act:
            Task.objects.filter(
                route=f"/orders/receiving/{order_id}/",
                assigned_to__role="storekeeper",
            ).exclude(status="done").update(status="done")
        observer = Employee.objects.filter(user=request.user, is_active=True).first()
        _create_manager_followup_task(
            order_id,
            latest.agency,
            request,
            timezone.localtime(),
            observer=observer,
        )
        return redirect(f"/orders/receiving/{order_id}/placement/?ok=1")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        order_id = kwargs.get("order_id")
        entries = self._load_entries(order_id)
        latest = entries[-1] if entries else None
        status_entry = _current_status_entry(entries)
        receiving_act = self._receiving_act_entry(entries)
        placement_act = self._placement_act_entry(entries)
        receiving_items = (receiving_act.payload or {}).get("act_items") if receiving_act else []
        placement_items = (placement_act.payload or {}).get("act_items") if placement_act else []
        display_items = []
        source_items = placement_items or receiving_items
        for item in source_items:
            display_items.append(
                {
                    "sku_code": item.get("sku_code") or "",
                    "name": item.get("name") or "-",
                    "size": item.get("size") or "-",
                    "actual_qty": item.get("actual_qty") or 0,
                    "box_qty": item.get("box_qty") or 0,
                    "pallet_qty": item.get("pallet_qty") or 0,
                }
            )
        catalog_items = []
        remaining_items = []
        for item in receiving_items:
            catalog_items.append(
                {
                    "sku_code": item.get("sku_code") or "",
                    "name": item.get("name") or "",
                    "size": item.get("size") or "",
                }
            )
            remaining_items.append(
                {
                    "sku_code": item.get("sku_code") or "",
                    "name": item.get("name") or "",
                    "size": item.get("size") or "",
                    "actual_qty": item.get("actual_qty") or 0,
                }
            )
        client_label = "-"
        if latest and latest.agency:
            name = latest.agency.agn_name or latest.agency.fio_agn or str(latest.agency)
            client_label = _shorten_ip_name(name)
        role = get_request_role(self.request)
        can_submit = role == "storekeeper"
        act_state = "open"
        if placement_act:
            act_state = (placement_act.payload or {}).get("act_state") or "closed"
        signed_by_storekeeper = _act_storekeeper_signed(entries)
        can_open_act = can_submit and act_state == "closed" and not signed_by_storekeeper
        boxes_data = (placement_act.payload or {}).get("act_boxes") if placement_act else []
        pallets_data = (placement_act.payload or {}).get("act_pallets") if placement_act else []
        ctx.update(
            {
                "order_id": order_id,
                "client_label": client_label,
                "status_label": _status_label_from_entry(status_entry) if status_entry else "-",
                "cabinet_url": resolve_cabinet_url(get_request_role(self.request)),
                "can_submit": can_submit,
                "can_open_act": can_open_act,
                "signed_by_storekeeper": signed_by_storekeeper,
                "act_exists": bool(placement_act),
                "act_state": act_state,
                "items": display_items,
                "catalog_items": catalog_items,
                "remaining_items": remaining_items,
                "boxes_data": boxes_data,
                "pallets_data": pallets_data,
                "ok": kwargs.get("ok", False),
                "error": kwargs.get("error"),
            }
        )
        return ctx
